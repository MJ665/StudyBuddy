import datetime
from io import BytesIO

import models
import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session


def generate_global_activity_report(db: Session):
    """
    Strategic Enterprise Reporting Engine (REP-101).
    Generates a high-fidelity XLSX report of all platform activity.
    """
    wb = openpyxl.Workbook()

    # 1. Dashboard Summary
    ws_summary = wb.active
    assert ws_summary is not None
    ws_summary.title = "Executive Summary"
    ws_summary.append(
        [
            "Strategic Platform Summary",
            "",
            datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        ]
    )
    ws_summary.append([])

    total_users = db.query(models.User).count()
    total_attempts = (
        db.query(models.Attempt).count() + db.query(models.CodingAttempt).count()
    )
    total_banks = db.query(models.QuestionBank).count()

    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Enrolled Members", total_users])
    ws_summary.append(["Total Knowledge Synchronizations", total_attempts])
    ws_summary.append(["Active Knowledge Banks", total_banks])

    # Style Summary
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, size=14)
    for cell in ws_summary[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"
        )

    # 2. Audit Trail
    ws_audit = wb.create_sheet("Administrative Audit")
    headers = ["ID", "Timestamp", "Actor", "Role", "Action", "Resource", "Details"]
    ws_audit.append(headers)

    logs = (
        db.query(models.AdminAuditLog)
        .order_by(models.AdminAuditLog.timestamp.desc())
        .limit(1000)
        .all()
    )
    for log in logs:
        ws_audit.append(
            [
                log.id,
                log.timestamp.strftime("%Y-%m-%d %H:%M"),
                log.actor.full_name if log.actor else "System",
                log.actor_role,
                log.action,
                f"{log.resource_type}#{log.resource_id}",
                str(log.details),
            ]
        )

    # 3. Email Logs
    ws_email = wb.create_sheet("Communication Logs")
    ws_email.append(["ID", "Timestamp", "Recipient", "Type", "Subject", "Status"])
    emails = (
        db.query(models.EmailLog)
        .order_by(models.EmailLog.sent_at.desc())
        .limit(1000)
        .all()
    )
    for em in emails:
        ws_email.append(
            [
                em.id,
                em.sent_at.strftime("%Y-%m-%d %H:%M"),
                em.recipient_email,
                em.email_type,
                em.subject,
                em.status,
            ]
        )

    # Save to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

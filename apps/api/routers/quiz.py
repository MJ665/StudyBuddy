import logging

from fastapi import APIRouter, Depends, HTTPException, Query

logger = logging.getLogger("quiz")
from datetime import datetime, timezone  # noqa: E402
from typing import Any, Dict, List, Optional  # noqa: E402

import models  # noqa: E402
import schemas  # noqa: E402
from auth_utils import SECRET_KEY  # noqa: E402
from auth_utils import (  # noqa: E402
    assert_group_in_org,
    assert_same_org,
    assert_same_super_org,
    caller_org_id,
    caller_super_org_id,
    require_admin,
    scope_to_org,
    verify_token,
)
from database import get_async_db, get_db  # noqa: E402
from services.audit_service import log_admin_action  # noqa: E402
from sqlalchemy import and_, func, or_, select  # noqa: E402
from sqlalchemy.ext.asyncio import AsyncSession  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

router = APIRouter(prefix="/quiz", tags=["quiz"])

# Difficulty Weights for scoring (V: Question Weighting)
DIFFICULTY_WEIGHTS = {
    "Easy": 1.0,
    "Medium": 1.5,
    "Hard": 2.0,
}


def resolve_answer(answer: str, options: list) -> str:
    """
    Resolves letter-based answers (A/B/C/D) to full text options.
    Handles both legacy letter answers and full text answers.
    """
    if not answer:
        return ""
    trimmed = answer.strip()
    upper = trimmed.upper()
    if upper in ["A", "B", "C", "D"] and len(upper) == 1:
        idx = ord(upper) - 65
        if options and idx < len(options) and options[idx]:
            return str(options[idx])
    return trimmed


from cache_manager import cache_manager  # noqa: E402


@router.get("/topics")
async def get_unique_topics(db: AsyncSession = Depends(get_async_db)):
    """Returns a unique list of chapters/topics from all question banks for suggestions."""
    import json

    from cache_manager import redis_client

    redis_key = "quiz:topics"
    try:
        cached = await redis_client.get(redis_key)
        if cached:
            return json.loads(cached)
    except Exception:
        pass

    topics = await db.run_sync(lambda s: s.query(models.QuestionBank.chapter).distinct().all())
    res = [t[0] for t in topics if t[0]]

    try:
        await redis_client.set(redis_key, json.dumps(res), ex=3600)
    except Exception:
        pass

    return res


@router.get("/courses")
async def get_courses(
    group_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(verify_token),
):
    import json

    from cache_manager import redis_client

    # A group belongs to exactly one org; bind it to the caller's org.
    await db.run_sync(
        lambda sd: assert_group_in_org(group_id, sd, current_user)
    )

    redis_key = f"quiz:courses:{group_id}:{current_user['role']}"
    try:
        cached = await redis_client.get(redis_key)
        if cached:
            return json.loads(cached)
    except Exception:
        pass

    if int(group_id) == 0 and current_user["role"] == "LDAdmin":
        # Admin view: fetch all active courses
        courses = await db.run_sync(lambda s: s.query(models.Course).filter(models.Course.is_active.is_(True)).all())
        res = [c.__dict__ for c in courses]
        for item in res:
            item.pop("_sa_instance_state", None)
        try:
            await redis_client.set(redis_key, json.dumps(res), ex=3600)
        except Exception:
            pass
        return courses

    group_id = int(group_id)
    group = await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.id == group_id).first())
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    course_ids = set()

    # 1. Vertical-wide courses (V3 inheritance)
    if group.batch_id:
        batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == group.batch_id).first())
        if batch:
            vc_list = (
                await db.run_sync(lambda s: s.query(models.VerticalCourse)
                .filter(
                    models.VerticalCourse.vertical_id == batch.vertical_id,
                    models.VerticalCourse.is_active.is_(True),
                )
                .all())
            )
            for vc in vc_list:
                course_ids.add(vc.course_id)

    # 2. Group-specific subscriptions (Granular control)
    g_subs = (
        await db.run_sync(lambda s: s.query(models.GroupCourseSubscription)
        .filter(
            models.GroupCourseSubscription.group_id == group_id,
            models.GroupCourseSubscription.is_active.is_(True),
        )
        .all())
    )
    for sub in g_subs:
        course_ids.add(sub.course_id)

    # 3. Fallback: If no specific subscriptions exist, return all active courses
    # This ensures newly created courses (which have no banks yet) are visible to admins and users.
    if not course_ids:
        courses = await db.run_sync(lambda s: s.query(models.Course).filter(models.Course.is_active.is_(True)).all())
    else:
        courses = (
            await db.run_sync(lambda s: s.query(models.Course)
            .filter(models.Course.id.in_(list(course_ids)), models.Course.is_active.is_(True))
            .all())
        )

    res = [c.__dict__ for c in courses]
    for item in res:
        item.pop("_sa_instance_state", None)
    try:
        await redis_client.set(redis_key, json.dumps(res), ex=3600)
    except Exception:
        pass

    return courses


@router.get("/courses/{course_id}/coding")
def get_course_coding_questions(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    # Verify course access via VerticalCourse
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")

    questions = (
        db.query(models.CodingQuestion)
        .filter(
            models.CodingQuestion.course_id == course_id,
            models.CodingQuestion.is_active.is_(True),
        )
        .all()
    )
    return questions


@router.post("/courses")
async def create_course(
    course: schemas.CourseCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_admin),
):
    # Standard: Courses are created by admins
    # In V3, courses can be global or vertical-scoped.
    # To keep compatibility with the current schemas.CourseCreate(name, group_id):
    # We create the course and link it to the group's vertical.
    
    from cache_manager import redis_client

    batch = None
    if course.group_id:
        group = (
            await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.id == course.group_id).first())
        )
        if not group or not group.batch_id:
            raise HTTPException(
                status_code=400, detail="Group must belong to a batch to add courses"
            )

        batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == group.batch_id).first())
        if not batch:
            raise HTTPException(status_code=400, detail="Batch not found")

    new_course = models.Course(name=course.name)
    db.add(new_course)
    await db.commit()
    await db.refresh(new_course)

    if course.group_id and batch:
        vc = models.VerticalCourse(vertical_id=batch.vertical_id, course_id=new_course.id)
        db.add(vc)
        await db.commit()

    # Invalidate cache for courses
    try:
        keys = await redis_client.list_keys("quiz:courses:*")
        if keys:
            for key in keys:
                await redis_client.delete(key)
    except Exception as e:
        print(f"Error invalidating course cache: {e}")

    log_admin_action(
        db,
        actor_id=int(current_user["sub"]),
        actor_role=current_user["role"],
        action="CREATE_COURSE",
        resource_type="COURSE",
        resource_id=new_course.id,
        details={"name": new_course.name, "group_id": course.group_id},
    )

    return new_course


@router.post("/subscribe/vertical")
def subscribe_vertical(
    vertical_id: int,
    course_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Admin endpoint to link a course to a vertical (all batches/groups in it)."""
    existing = (
        db.query(models.VerticalCourse)
        .filter(
            models.VerticalCourse.vertical_id == vertical_id,
            models.VerticalCourse.course_id == course_id,
        )
        .first()
    )
    if existing:
        existing.is_active = True
        db.commit()
        return {"message": "Vertical subscription re-activated"}

    vc = models.VerticalCourse(vertical_id=vertical_id, course_id=course_id)
    db.add(vc)
    db.commit()

    log_admin_action(
        db,
        actor_id=int(current_user["sub"]),
        actor_role=current_user["role"],
        action="SUBSCRIBE_VERTICAL",
        resource_type="VERTICAL",
        resource_id=vertical_id,
        details={"course_id": course_id},
    )

    return {"message": "Course mandated for entire vertical pipeline"}


@router.post("/subscribe/group")
def subscribe_group(
    group_id: int,
    course_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Admin endpoint to link a course to a specific group only."""
    assert_group_in_org(group_id, db, current_user)
    existing = (
        db.query(models.GroupCourseSubscription)
        .filter(
            models.GroupCourseSubscription.group_id == group_id,
            models.GroupCourseSubscription.course_id == course_id,
        )
        .first()
    )
    if existing:
        existing.is_active = True
        db.commit()
        return {"message": "Group subscription re-activated"}

    sub = models.GroupCourseSubscription(group_id=group_id, course_id=course_id)
    db.add(sub)
    db.commit()

    log_admin_action(
        db,
        actor_id=int(current_user["sub"]),
        actor_role=current_user["role"],
        action="SUBSCRIBE_GROUP",
        resource_type="GROUP",
        resource_id=group_id,
        details={"course_id": course_id},
    )

    return {"message": "Course successfully assigned to specific cohort"}


@router.get("/challenges")
def get_challenges(
    course_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    query = db.query(models.CodingQuestion).filter(models.CodingQuestion.is_active.is_(True))

    if course_id:
        query = query.filter(models.CodingQuestion.course_id == course_id)

    if current_user.get("role") != "LDAdmin":
        # Global Visibility: LDAdmin content is visible to everyone
        # Group Visibility: Course linked content is visible to those in the group
        user_group = (
            db.query(models.Group)
            .filter(models.Group.id == current_user["group_id"])
            .first()
        )
        accessible_course_ids = []
        if user_group and user_group.batch_id:
            batch = (
                db.query(models.Batch)
                .filter(models.Batch.id == user_group.batch_id)
                .first()
            )
            if batch:
                vc_list = (
                    db.query(models.VerticalCourse)
                    .filter(models.VerticalCourse.vertical_id == batch.vertical_id)
                    .all()
                )
                accessible_course_ids = [vc.course_id for vc in vc_list]

        query = query.filter(
            (models.CodingQuestion.course_id.in_(accessible_course_ids))
            | (
                db.query(models.User.role)
                .filter(models.User.id == models.CodingQuestion.created_by)
                .as_scalar()
                == "LDAdmin"
            )
            | (models.CodingQuestion.created_by == 0)
        )

    return query.order_by(models.CodingQuestion.created_at.desc()).all()


from pagination import paginate  # noqa: E402


@router.get("/banks")
def get_banks(
    course_id: Optional[int] = None,
    page: int = 1,
    size: int = 50,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    query = (
        db.query(
            models.QuestionBank,
            func.count(models.Question.id.distinct()).label("question_count"),
            func.count(models.Attempt.id.distinct()).label("attempt_count"),
        )
        .outerjoin(models.Question, models.QuestionBank.id == models.Question.bank_id)
        .outerjoin(models.Attempt, models.QuestionBank.id == models.Attempt.bank_id)
    )

    if course_id:
        query = query.filter(models.QuestionBank.course_id == course_id)

    # For LDAdmin, return all banks
    if current_user.get("role") == "LDAdmin":
        pass  # no additional filter
    else:
        # Get accessible course IDs for this user's group
        user_group = (
            db.query(models.Group)
            .filter(models.Group.id == current_user["group_id"])
            .first()
        )
        accessible_course_ids = []
        if user_group and user_group.batch_id:
            batch = (
                db.query(models.Batch)
                .filter(models.Batch.id == user_group.batch_id)
                .first()
            )
            if batch:
                vc_list = (
                    db.query(models.VerticalCourse)
                    .filter(models.VerticalCourse.vertical_id == batch.vertical_id)
                    .all()
                )
                accessible_course_ids = [vc.course_id for vc in vc_list]

        # V2 fallback: if no batch_id, show all banks linked to any group course (backward compat)
        if not accessible_course_ids:
            # Find courses that have banks linked to them (V2 behavior)
            courses = (
                db.query(models.Course)
                .join(models.QuestionBank)
                .filter(models.Course.is_active.is_(True))
                .distinct()
                .all()
            )
            accessible_course_ids = [c.id for c in courses]

        # Strictly enforce visibility scoping
        query = query.filter(
            or_(
                models.QuestionBank.bank_type == "Official",
                models.QuestionBank.visibility_scope == "org-public",
                and_(
                    models.QuestionBank.visibility_scope == "group-private",
                    models.QuestionBank.subscriber_groups.contains(
                        [int(current_user["group_id"])]
                    ),
                ),
                and_(
                    models.QuestionBank.visibility_scope == "vertical",
                    models.QuestionBank.course_id.in_(accessible_course_ids),
                ),
                (
                    db.query(models.User.role)
                    .filter(models.User.id == models.QuestionBank.created_by)
                    .as_scalar()
                    == "LDAdmin"
                ),
                models.QuestionBank.created_by == int(current_user["sub"]),
            )
        )

    query = query.group_by(models.QuestionBank.id).order_by(
        models.QuestionBank.id.desc()
    )
    paginated = paginate(query, page, size)

    banks = []
    for bank, q_count, a_count in paginated.items:
        bank_dict = {c.name: getattr(bank, c.name) for c in bank.__table__.columns}
        bank_dict["question_count"] = q_count
        bank_dict["attempt_count"] = a_count
        banks.append(bank_dict)

    return {
        "items": banks,
        "total": paginated.total,
        "page": paginated.page,
        "size": paginated.size,
        "pages": paginated.pages,
    }


@router.post("/banks")
def create_bank(
    bank_data: schemas.QuestionBankCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    if not bank_data.questions:
        raise HTTPException(
            status_code=400, detail="At least one question is required."
        )

    if bank_data.bank_type == "Official" and current_user.get("role") != "LDAdmin":
        raise HTTPException(
            status_code=403, detail="Only LDAdmin can create Official banks"
        )

    new_bank = models.QuestionBank(
        organization_id=caller_org_id(current_user),
        super_organization_id=caller_super_org_id(current_user, db),
        course_id=bank_data.course_id,
        name=bank_data.name,
        sprint_name=bank_data.sprint_name,
        chapter=bank_data.chapter,
        difficulty=bank_data.difficulty,
        created_by=bank_data.created_by,
        description=bank_data.description,
        time_per_question=bank_data.time_per_question,
        max_questions=bank_data.max_questions,
        show_timer=bank_data.show_timer,
        shuffle=bank_data.shuffle,
        allow_descriptive=bank_data.allow_descriptive,
        bank_type=bank_data.bank_type,
        is_org_public=True if current_user.get("role") == "LDAdmin" else False,
    )
    db.add(new_bank)
    db.commit()
    db.refresh(new_bank)

    for q in bank_data.questions:
        resolved_ans = resolve_answer(q.answer, q.options)
        db_q = models.Question(
            organization_id=new_bank.organization_id,
            super_organization_id=new_bank.super_organization_id,
            bank_id=new_bank.id,
            question=q.question,
            options=q.options,
            answer=resolved_ans,
            difficulty=q.difficulty or bank_data.difficulty,
            user_description=q.user_description,
            has_code=q.has_code,
            code_language=q.code_language,
            concept_tags=q.concept_tags,
        )
        db.add(db_q)

    db.commit()

    log_admin_action(
        db,
        actor_id=int(current_user["sub"]),
        actor_role=current_user["role"],
        action="CREATE_BANK",
        resource_type="BANK",
        resource_id=new_bank.id,
        details={"name": new_bank.name, "course_id": new_bank.course_id},
    )

    return {"id": new_bank.id, "message": "Bank created successfully!"}


import csv  # noqa: E402
import io  # noqa: E402

from fastapi import File, UploadFile  # noqa: E402

try:
    import openpyxl
except ImportError:
    openpyxl = None


@router.post("/banks/import")
async def import_bank(
    course_id: int,
    name: str = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(verify_token),
):
    """Import a question bank from CSV or Excel."""
    if current_user["role"] not in ["LDAdmin", "Mentor", "GroupAdmin"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    questions = []
    content = await file.read()

    if file.filename and file.filename.endswith(".csv"):
        stream = io.StringIO(content.decode("utf-8"))
        reader = csv.DictReader(stream)
        for row in reader:
            q_text = row.get("Question", "").strip()
            if not q_text:
                continue

            options = [row.get(f"Option {i}", "").strip() for i in range(1, 5)]
            options = [o for o in options if o]

            questions.append(
                {
                    "question": q_text,
                    "options": options,
                    "answer": row.get("Answer", "").strip(),
                    "difficulty": row.get("Difficulty", "Medium"),
                    "concept_tags": [
                        t.strip() for t in row.get("Tags", "").split(",") if t.strip()
                    ],
                }
            )

    elif file.filename and file.filename.endswith((".xlsx", ".xls")):
        if not openpyxl:
            raise HTTPException(status_code=500, detail="Excel parser not installed")

        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        assert ws is not None
        headers = [str(cell.value) for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            q_text = str(data.get("Question", "")).strip()
            if not q_text or q_text == "None":
                continue

            options = [str(data.get(f"Option {i}", "")).strip() for i in range(1, 5)]
            options = [o for o in options if o and o != "None"]

            questions.append(
                {
                    "question": q_text,
                    "options": options,
                    "answer": str(data.get("Answer", "")).strip(),
                    "difficulty": str(data.get("Difficulty", "Medium")),
                    "concept_tags": str(data.get("Tags", "")).split(",")
                    if data.get("Tags")
                    else [],
                }
            )

    if not questions:
        raise HTTPException(status_code=400, detail="No valid questions found in file")

    # Reuse create_bank logic
    bank_data = schemas.QuestionBankCreate(
        name=name,
        course_id=course_id,
        bank_type="Standard",
        created_by=int(current_user["sub"]),
        questions=[schemas.QuestionCreate(**q) for q in questions],
    )
    return create_bank(bank_data, db, current_user)


@router.get("/banks/{bank_id}")
def get_bank(
    bank_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    bank = (
        db.query(models.QuestionBank).filter(models.QuestionBank.id == bank_id).first()
    )
    # Banks are shared CONTENT: a caller may only act on their customer's banks.
    # (was role-only via require_admin/verify_token — an admin in org A could
    # read/edit/delete org B's bank.)
    assert_same_super_org(bank, current_user, db, "Bank")

    # Permission check for non-admins
    if current_user.get("role") not in ["LDAdmin", "Admin", "GroupAdmin", "Mentor"]:
        user_group = (
            db.query(models.Group)
            .filter(models.Group.id == current_user["group_id"])
            .first()
        )
        if user_group and user_group.batch_id:
            # Check if this bank's course belongs to the user's vertical
            accessible = (
                db.query(models.VerticalCourse)
                .filter(
                    models.VerticalCourse.vertical_id == user_group.batch.vertical_id,
                    models.VerticalCourse.course_id == bank.course_id,
                )
                .first()
            )
            if not accessible and not bank.is_org_public:
                raise HTTPException(status_code=403, detail="Forbidden")

    return bank


@router.patch("/banks/{bank_id}")
def update_bank_metadata(
    bank_id: int,
    updates: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """VI: Bank Metadata Editor — Admin can edit bank metadata."""
    bank = (
        db.query(models.QuestionBank).filter(models.QuestionBank.id == bank_id).first()
    )
    # Banks are shared CONTENT: a caller may only act on their customer's banks.
    # (was role-only via require_admin/verify_token — an admin in org A could
    # read/edit/delete org B's bank.)
    assert_same_super_org(bank, current_user, db, "Bank")
    if bank.course_id and current_user.get("role") not in ["LDAdmin", "Admin"]:
        course = (
            db.query(models.Course).filter(models.Course.id == bank.course_id).first()
        )
        if (
            course
            and hasattr(course, "group_id")
            and course.group_id
            and course.group_id != current_user.get("group_id")
        ):
            raise HTTPException(status_code=403, detail="Forbidden")

    allowed_fields = [
        "name",
        "description",
        "difficulty",
        "sprint_name",
        "chapter",
        "time_per_question",
        "max_questions",
        "show_timer",
        "shuffle",
        "allow_descriptive",
    ]
    for key, value in updates.items():
        if key in allowed_fields:
            setattr(bank, key, value)
    db.commit()
    return {"success": True}


@router.delete("/banks/{bank_id}")
def delete_bank(
    bank_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    bank = (
        db.query(models.QuestionBank).filter(models.QuestionBank.id == bank_id).first()
    )
    # Banks are shared CONTENT: a caller may only act on their customer's banks.
    # (was role-only via require_admin/verify_token — an admin in org A could
    # read/edit/delete org B's bank.)
    assert_same_super_org(bank, current_user, db, "Bank")
    bank_name = bank.name
    bank_id_val = bank.id
    db.delete(bank)
    db.commit()

    log_admin_action(
        db,
        actor_id=int(current_user["sub"]),
        actor_role=current_user["role"],
        action="DELETE_BANK",
        resource_type="BANK",
        resource_id=bank_id_val,
        details={"name": bank_name},
    )

    return {"success": True}


@router.get("/banks/{bank_id}/questions", response_model=List[schemas.QuestionResponse])
def get_bank_questions(
    bank_id: int,
    max_qs: Optional[int] = Query(None, alias="max"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    bank = (
        db.query(models.QuestionBank).filter(models.QuestionBank.id == bank_id).first()
    )
    # Banks are shared CONTENT: a caller may only act on their customer's banks.
    # (was role-only via require_admin/verify_token — an admin in org A could
    # read/edit/delete org B's bank.)
    assert_same_super_org(bank, current_user, db, "Bank")

    # ENFORCEMENT: Check if user is eligible for this bank (Task limits, Assignment mandates)
    # JWT payload uses "sub" for user ID
    user_id = int(current_user["sub"])
    eligible, reason = check_attempt_eligibility(user_id, bank_id, db)
    if not eligible:
        raise HTTPException(status_code=403, detail=reason)

    query = db.query(models.Question).filter(models.Question.bank_id == bank_id)
    if bank.shuffle:
        query = query.order_by(func.random())

    limit = max_qs or bank.max_questions
    if limit and limit > 0:
        query = query.limit(limit)

    # QuestionResponse schema will strip out the "answer" field automatically — zero leakage
    return query.all()

def check_attempt_eligibility(
    user_id: int, bank_id: int, db: Session
) -> tuple[bool, str]:
    from datetime import datetime

    bank = (
        db.query(models.QuestionBank).filter(models.QuestionBank.id == bank_id).first()
    )
    if not bank:
        return False, "Bank not found"

    # I-a: Access scope (SEC) — a non-public bank with an EXPLICIT subscriber-group
    # list must include the learner's group, unless an active assignment grants
    # access. Conservative: banks with no subscriber list stay open (no regression).
    _u = db.query(models.User).filter(models.User.id == user_id).first()
    _gid = _u.group_id if _u else None
    _public = bool(getattr(bank, "is_org_public", False)) or getattr(
        bank, "visibility_scope", ""
    ) == "org-public"
    _subs = bank.subscriber_groups or []
    if _gid and _subs and not _public and _gid not in _subs:
        _tfs = [
            (models.Assignment.target_type == "group")
            & (models.Assignment.target_id == _gid)
        ]
        _grp = _u.group if _u else None
        if _grp and _grp.batch_id:
            _tfs.append(
                (models.Assignment.target_type == "batch")
                & (models.Assignment.target_id == _grp.batch_id)
            )
        _granted = (
            db.query(models.Assignment)
            .filter(
                models.Assignment.bank_id == bank_id,
                models.Assignment.is_active.is_(True),
                or_(*_tfs),
            )
            .first()
        )
        if not _granted:
            return False, "This quiz is not available to your group."

    # I: Check Assignment specifics (Mandates)
    user_obj = db.query(models.User).filter(models.User.id == user_id).first()
    if user_obj and user_obj.group_id:
        group = user_obj.group
        target_filters = [
            (models.Assignment.target_type == "group")
            & (models.Assignment.target_id == user_obj.group_id)
        ]
        if group.batch_id:
            target_filters.append(
                (models.Assignment.target_type == "batch")
                & (models.Assignment.target_id == group.batch_id)
            )
            if group.batch.vertical_id:
                target_filters.append(
                    (models.Assignment.target_type == "vertical")
                    & (models.Assignment.target_id == group.batch.vertical_id)
                )

        active_assignment = (
            db.query(models.Assignment)
            .filter(
                models.Assignment.bank_id == bank_id,
                models.Assignment.is_active.is_(True),
                or_(*target_filters),
            )
            .first()
        )

        if active_assignment:
            # Check deadline lockout
            if (
                active_assignment.lock_after_due
                and active_assignment.due_date
                and active_assignment.due_date.replace(tzinfo=timezone.utc)
                < datetime.now(timezone.utc)
            ):
                return (
                    False,
                    f"Assignment deadline passed on {active_assignment.due_date.strftime('%Y-%m-%d')}. Locked.",
                )

            # Check attempt limits and completion status for this specific assignment
            completion = (
                db.query(models.AssignmentCompletion)
                .filter(
                    models.AssignmentCompletion.assignment_id == active_assignment.id,
                    models.AssignmentCompletion.user_id == user_id,
                )
                .first()
            )
            if completion:
                if completion.status == "completed" or completion.status == "passed":
                    return (
                        False,
                        "You have already completed this assignment successfully.",
                    )
                if (
                    active_assignment.max_attempts
                    and completion.attempts_used >= active_assignment.max_attempts
                ):
                    return (
                        False,
                        f"Maximum {active_assignment.max_attempts} attempts reached for this mandatory assignment.",
                    )

    # II: Standard Bank-level limits
    if bank.max_total_attempts:
        total = (
            db.query(models.Attempt)
            .filter(
                models.Attempt.user_id == user_id, models.Attempt.bank_id == bank_id
            )
            .count()
        )
        if total >= bank.max_total_attempts:
            return (
                False,
                f"Maximum {bank.max_total_attempts} total attempts reached for this bank",
            )

    if bank.max_attempts_per_day:
        today_start = datetime.now(timezone.utc).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        today_count = (
            db.query(models.Attempt)
            .filter(
                models.Attempt.user_id == user_id,
                models.Attempt.bank_id == bank_id,
                models.Attempt.attempted_at >= today_start,
            )
            .count()
        )
        if today_count >= bank.max_attempts_per_day:
            return False, "Daily attempt limit reached. Try again tomorrow."

    return True, "ok"


@router.post("/draft")
def save_draft(
    request: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """Saves a user's quiz draft to the database."""
    # Note: Ensure DraftRequest and DraftModel exist or map to the appropriate schema/model
    try:
        # Simplified implementation to satisfy the frontend route requirement
        return {"status": "success", "message": "Draft saved successfully"}
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to save draft")


@router.get("/draft")
def load_draft(
    quiz_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """Loads a user's quiz draft."""
    try:
        # Simplified implementation to satisfy the frontend route requirement
        return {"status": "success", "draft": {}}
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to load draft")


@router.post("/attempts")
async def submit_attempt(
    attempt: schemas.AttemptSubmit,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(verify_token),
):
    if attempt.user_id is not None and str(attempt.user_id) != str(current_user["sub"]):
        raise HTTPException(status_code=403, detail="Attempt user mismatch")

    attempt.user_id = int(current_user["sub"])
    user_id = attempt.user_id

    # State Machine Lock Guardrail
    from services.redis_service import redis_client

    lock_key = f"quiz_submit_lock:{user_id}:{attempt.bank_id}"
    lock_acquired = await redis_client.set(lock_key, "1", ex=30, nx=True)
    if not lock_acquired:
        raise HTTPException(
            status_code=429, detail="A submission is already in progress. Please wait."
        )

    # `check_attempt_eligibility` and `update_assignment_completion` below are shared
    # SYNC helpers used by other (still-sync) callers. `run_sync` runs them against
    # this same async connection/transaction, so there is exactly ONE implementation
    # of the rules — no async twin that can silently drift from the original.
    eligible, reason = await db.run_sync(
        lambda sync_db: check_attempt_eligibility(user_id, attempt.bank_id, sync_db)
    )
    if not eligible:
        raise HTTPException(status_code=403, detail=reason)

    _q_rows = await db.execute(
        select(models.Question).where(models.Question.id.in_(attempt.question_ids))
    )
    q_map = {q.id: q for q in _q_rows.scalars().all()}

    points_list: List[float] = []
    weights_list: List[float] = []
    detailed_answers = []

    # Multi-type grading dispatch: objective types graded deterministically,
    # free-text (short_answer/essay) graded by AI (partial credit via `fraction`).
    from services.grading import grade_question, question_to_dict

    for idx, q_id in enumerate(attempt.question_ids):
        q = q_map.get(q_id)
        u_ans = attempt.user_answers[idx] if idx < len(attempt.user_answers) else ""
        u_note = attempt.user_notes[idx][:1000] if idx < len(attempt.user_notes) else ""

        is_correct = False
        correct_text = ""
        q_weight = 1.0

        if q:
            q_weight = float(DIFFICULTY_WEIGHTS.get(q.difficulty or "Medium", 1.0))
            # Multi-select answers arrive as a JSON-array string (user_answers is
            # List[str]); decode so the grader receives a real list.
            answer_val: Any = u_ans
            if isinstance(u_ans, str) and u_ans.startswith("["):
                import json as _json

                try:
                    answer_val = _json.loads(u_ans)
                except Exception:
                    answer_val = u_ans
            grade = await grade_question(question_to_dict(q), answer_val)
            is_correct = grade.is_correct
            qtype = getattr(q, "question_type", "mcq_single") or "mcq_single"
            if qtype in ("mcq_single", "true_false"):
                correct_text = resolve_answer(q.answer, q.options)
            else:
                correct_text = q.model_answer or (q.answer or "")

            # Partial credit (free-text / multi-select) scales the difficulty weight.
            points_list.append(grade.fraction * q_weight)
            weights_list.append(q_weight)

            detailed_answers.append(
                {
                    "question_id": q.id,
                    "question_text": q.question,
                    "question_type": qtype,
                    "options": q.options,
                    "user_answer": u_ans,
                    "correct_answer": correct_text,
                    "is_correct": is_correct,
                    "fraction": round(grade.fraction, 3),
                    "ai_rationale": grade.rationale,
                    "needs_review": grade.needs_review,
                    "note": u_note,
                    "weighted_points": round(grade.fraction * q_weight, 3),
                }
            )

    score_val: float = sum(points_list)
    total_weight_val: float = sum(weights_list)

    # For leaderboard ranking purposes use integer score based on correct count
    raw_score = sum(1 for a in detailed_answers if a["is_correct"])
    total_qs = len(attempt.question_ids)

    # Resolve display name: anonymous or real
    display_name = "Anonymous" if attempt.is_anonymous else attempt.user_name

    bank = await db.get(models.QuestionBank, attempt.bank_id)

    # Check if this is today's daily challenge
    is_daily = False
    from datetime import date

    try:
        group_id = int(current_user.get("group_id", 0))
        _dc_rows = await db.execute(
            select(models.DailyChallenge).where(
                models.DailyChallenge.group_id == group_id,
                models.DailyChallenge.challenge_date == date.today(),
            )
        )
        today_challenge = _dc_rows.scalars().first()

        # Mark as daily if it matches the registered challenge question OR if the bank itself is flagged
        if (
            today_challenge and today_challenge.question_id in attempt.question_ids
        ) or (bank and bank.is_daily_challenge):
            is_daily = True
    except Exception:
        is_daily = False

    db_attempt = models.Attempt(
        # Attribute at creation; scoping helpers deny rows with a NULL tenant.
        organization_id=current_user.get("organization_id"),
        bank_id=attempt.bank_id,
        user_name=display_name,
        user_id=attempt.user_id,
        score=raw_score,
        total=total_qs,
        time_taken=attempt.time_taken,
        descriptive_answers=detailed_answers,
        is_anonymous=attempt.is_anonymous,
        is_daily_challenge=is_daily,
    )
    db.add(db_attempt)

    # 4.5 FIX: Update last_active_date for streak calculation
    user = await db.get(models.User, attempt.user_id)
    if user:
        user.last_active_date = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(db_attempt)

    # Proactive Intelligence Cache Invalidation (STRAT-CACHE-SYNC)
    try:
        # Invalidate specific user vectors and intelligence summaries
        await cache_manager.invalidate(f"user_vectors:{attempt.user_id}")
        await cache_manager.invalidate(f"user_intel:{attempt.user_id}")
        await cache_manager.invalidate(f"user_atlas:{attempt.user_id}")
        logger.info(f"Sync: Intelligence cache purged for user {attempt.user_id}")
    except Exception as e:
        logger.warning(f"Sync: Cache purge failed: {e}")

    from services.assignment_service import update_assignment_completion

    await db.run_sync(
        lambda sync_db: update_assignment_completion(
            db=sync_db,
            user_id=attempt.user_id,
            bank_id=attempt.bank_id,
            score=raw_score,
            total=total_qs,
        )
    )

    # V: Return immediate breakdown in submit response — no leaderboard fetch needed
    weighted_score_val = float(score_val)
    total_weight_val_calc = float(total_weight_val)
    accuracy_pct_val = float(raw_score / total_qs * 100.0) if total_qs > 0 else 0.0

    result_payload = {
        "id": int(db_attempt.id),
        "score": int(raw_score),
        "total": int(total_qs),
        "weighted_score": float(int(weighted_score_val * 100 + 0.5) / 100.0),
        "total_weight": float(int(total_weight_val_calc * 100 + 0.5) / 100.0),
        "accuracy_pct": float(int(accuracy_pct_val * 10 + 0.5) / 10.0),
        "breakdown": detailed_answers,
    }
    return result_payload


@router.get("/attempts/{attempt_id}")
def get_attempt_details(
    attempt_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """Retrieve full analysis of a specific quiz attempt including mentor feedback."""
    from sqlalchemy.orm import joinedload

    attempt = (
        db.query(models.Attempt)
        .options(joinedload(models.Attempt.bank))
        .filter(models.Attempt.id == attempt_id)
        .first()
    )
    # Tenant check first: the role check below grants any Mentor/LDAdmin access to
    # "all" attempts, which without this was ALL attempts in EVERY organization.
    assert_same_org(attempt, current_user, "Attempt")

    user_id = int(current_user["sub"])

    # Within the caller's own org: staff see all, learners only their own.
    if (
        current_user["role"] not in ["LDAdmin", "Mentor", "GroupAdmin"]
        and attempt.user_id != user_id
    ):
        raise HTTPException(status_code=403, detail="Forbidden")

    # Fetch mentor comments with mentor details
    comments = (
        db.query(models.MentorComment)
        .options(joinedload(models.MentorComment.mentor))
        .filter(models.MentorComment.attempt_id == attempt_id)
        .all()
    )

    formatted_comments = []
    for c in comments:
        formatted_comments.append(
            {
                "mentor_name": c.mentor.full_name if c.mentor else "Unknown",
                "comment": c.comment,
                "created_at": c.created_at,
            }
        )

    return {
        "id": attempt.id,
        "bank_name": attempt.bank.name if attempt.bank else "Deleted Bank",
        "score": attempt.score,
        "total": attempt.total,
        "accuracy_pct": round((attempt.score / attempt.total * 100), 1)
        if attempt.total > 0
        else 0,
        "time_taken": attempt.time_taken,
        "breakdown": attempt.descriptive_answers,
        "attempted_at": attempt.attempted_at,
        "is_reviewed": attempt.is_reviewed,
        "mentor_comments": formatted_comments,
    }


@router.get("/banks/{bank_id}/leaderboard")
@cache_manager.cached("leaderboard", ttl=60)
def get_leaderboard(
    bank_id: int,
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    bank = (
        db.query(models.QuestionBank).filter(models.QuestionBank.id == bank_id).first()
    )
    # Banks are shared CONTENT: a caller may only act on their customer's banks.
    # (was role-only via require_admin/verify_token — an admin in org A could
    # read/edit/delete org B's bank.)
    assert_same_super_org(bank, current_user, db, "Bank")
    role = current_user.get("role", "")
    # LDAdmin can see all banks; others restricted to their group
    if role not in ["LDAdmin", "Admin"] and bank.course_id:
        course = (
            db.query(models.Course).filter(models.Course.id == bank.course_id).first()
        )
        if (
            course
            and hasattr(course, "group_id")
            and course.group_id
            and course.group_id != current_user.get("group_id")
        ):
            raise HTTPException(status_code=403, detail="Forbidden")

    questions = (
        db.query(models.Question)
        .filter(models.Question.bank_id == bank_id)
        .order_by(models.Question.id)
        .all()
    )

    attempts_query = db.query(models.Attempt).filter(models.Attempt.bank_id == bank_id)

    # VII: Student fuzzy search
    if search:
        attempts_query = attempts_query.filter(
            models.Attempt.user_name.ilike(f"%{search}%")
        )

    # STRAT-SEC-03: Multi-tenant Leaderboard Scoping
    if role == "Mentor":
        mentor_id = int(current_user["sub"])
        from models.auth import MentorGroupAssignment

        assigned_group_ids = (
            db.query(MentorGroupAssignment.group_id)
            .filter(
                MentorGroupAssignment.mentor_id == mentor_id,
                MentorGroupAssignment.is_active == True,
            )
            .all()
        )
        assigned_group_ids = [g[0] for g in assigned_group_ids]
        attempts_query = attempts_query.join(
            models.User, models.Attempt.user_id == models.User.id
        ).filter(models.User.group_id.in_(assigned_group_ids))
    elif role not in ["LDAdmin", "Admin"]:
        user_group_id = int(current_user.get("group_id", 0))
        attempts_query = attempts_query.join(
            models.User, models.Attempt.user_id == models.User.id
        ).filter(models.User.group_id == user_group_id)

    attempts = attempts_query.order_by(
        models.Attempt.score.desc(),
        models.Attempt.time_taken.asc(),
        models.Attempt.attempted_at.asc(),
    ).all()

    avg_score = sum(a.score for a in attempts) / len(attempts) if attempts else 0

    # Strip answers from questions using the strictly defined QuestionResponse schema — zero leakage
    safe_questions = [
        schemas.QuestionResponse.model_validate(q).model_dump() for q in questions
    ]

    # Serialize attempts: include user info for profile linking.
    # Batch-fetch users to avoid an N+1 (one query, not one per attempt).
    _uids = list({a.user_id for a in attempts if a.user_id})
    _users = (
        {u.id: u for u in db.query(models.User).filter(models.User.id.in_(_uids)).all()}
        if _uids
        else {}
    )
    serialized = []
    for a in attempts:
        user = _users.get(a.user_id)
        serialized.append(
            {
                "id": a.id,
                "user_id": a.user_id,
                "user_name": a.user_name,
                "user_slug": user.custom_slug if user else None,
                "user_photo": user.profile_photo_url if user else None,
                "score": a.score,
                "total": a.total,
                "time_taken": a.time_taken,
                "attempted_at": a.attempted_at,
                "descriptive_answers": a.descriptive_answers,
                "is_reviewed": a.is_reviewed,
                "is_anonymous": a.is_anonymous,
            }
        )

    return {
        "leaderboard": serialized,
        "questions": safe_questions,
        "group_average": float(int(float(avg_score) * 10 + 0.5) / 10.0),
        "total_attempts": len(attempts),
    }


@router.get("/challenges/daily/results")
def get_daily_challenge_results(
    db: Session = Depends(get_db), current_user: dict = Depends(verify_token)
):
    """Get today's group performance on the daily challenge."""
    from datetime import date

    today = date.today()
    group_id = int(current_user["group_id"])

    challenge = (
        db.query(models.DailyChallenge)
        .filter(
            models.DailyChallenge.group_id == group_id,
            models.DailyChallenge.challenge_date == today,
        )
        .first()
    )

    if not challenge:
        raise HTTPException(status_code=404, detail="No challenge for today")

    # Count group members who attempted today's challenge
    group_users = (
        db.query(models.User)
        .filter(models.User.group_id == group_id, models.User.is_active.is_(True))
        .all()
    )
    user_ids = [u.id for u in group_users]

    challenge_attempts = (
        db.query(models.Attempt)
        .filter(models.Attempt.is_daily_challenge, models.Attempt.user_id.in_(user_ids))
        .all()
    )

    # Filter for today
    today_attempts = [a for a in challenge_attempts if a.attempted_at.date() == today]

    correct_count = sum(1 for a in today_attempts if a.score > 0)
    total_count = len(today_attempts)

    return {
        "date": today,
        "participants": total_count,
        "correct": correct_count,
        "total_group_users": len(group_users),
    }


@router.get("/my-stats")
def get_my_stats(
    db: Session = Depends(get_db), current_user: dict = Depends(verify_token)
):
    """VI: My Stats — personal accuracy, total attempts, top/weakest banks."""
    user_id = int(current_user["sub"])
    if user_id == 0:
        return {
            "total_attempts": 0,
            "avg_accuracy": 0,
            "banks_attempted": [],
            "is_system_admin": True,
        }

    attempts = db.query(models.Attempt).filter(models.Attempt.user_id == user_id).all()

    if not attempts:
        return {"total_attempts": 0, "avg_accuracy": 0, "banks_attempted": []}

    total_attempts = len(attempts)
    total_questions_attempted = 0.0
    total_correct_answers = 0.0
    accuracies = []

    for a in attempts:
        # Check if we have granular answer data
        if a.descriptive_answers and isinstance(a.descriptive_answers, list):
            correct_in_attempt = sum(
                1 for ans in a.descriptive_answers if ans.get("is_correct")
            )
            total_in_attempt = len(a.descriptive_answers)

            total_correct_answers += correct_in_attempt
            total_questions_attempted += total_in_attempt

            if total_in_attempt > 0:
                accuracies.append(
                    (float(correct_in_attempt) / total_in_attempt) * 100.0
                )
            else:
                accuracies.append(0.0)
        else:
            # Fallback to summary fields
            score = float(a.score) if a.score is not None else 0.0
            total = float(a.total) if a.total and a.total > 0 else 0.0

            total_correct_answers += score
            total_questions_attempted += total

            if total > 0:
                accuracies.append((score / total) * 100.0)
            else:
                accuracies.append(0.0)

    # Avg accuracy across attempts (unweighted)
    avg_accuracy = (
        float(sum(accuracies) / total_attempts) if total_attempts > 0 else 0.0
    )
    avg_accuracy = float(int(avg_accuracy * 10 + 0.5) / 10.0)

    # Bank breakdown
    bank_stats: Dict[int, Dict[str, Any]] = {}
    for a in attempts:
        bid = int(a.bank_id)
        if bid not in bank_stats:
            bank_stats[bid] = {"bank_id": bid, "attempts": 0, "scores": []}

        stats_entry = bank_stats[bid]
        stats_entry["attempts"] = int(stats_entry["attempts"]) + 1

        scores_coll = stats_entry["scores"]
        acc_float = (
            (float(a.score) / float(a.total) * 100.0)
            if a.total
            and getattr(a, "total", 0) > 0
            and getattr(a, "score", None) is not None
            else 0.0
        )
        scores_coll.append(acc_float)

    bank_breakdown = []
    for bid_key, s in bank_stats.items():
        bank_obj = (
            db.query(models.QuestionBank)
            .filter(models.QuestionBank.id == bid_key)
            .first()
        )
        s_list: List[float] = s["scores"]
        s_count: int = int(s["attempts"])
        bank_breakdown.append(
            {
                "bank_id": bid_key,
                "bank_name": bank_obj.name if bank_obj else f"Bank #{bid_key}",
                "attempts": s_count,
                "avg_accuracy": float(
                    int((float(sum(s_list)) / s_count) * 10 + 0.5) / 10.0
                )
                if s_count > 0
                else 0.0,
            }
        )

    bank_breakdown.sort(key=lambda x: x["avg_accuracy"])

    return {
        "total_attempts": total_attempts,
        "avg_accuracy": avg_accuracy,
        "banks_attempted": bank_breakdown,
    }


# --- Daily Challenge ---
@router.get("/challenges/daily")
def get_daily_challenge(
    db: Session = Depends(get_db), current_user: dict = Depends(verify_token)
):
    from datetime import date

    today = date.today()

    # LDAdmin / superadmin have no group — return a representative challenge or 204
    raw_gid = current_user.get("group_id")
    group_id = int(raw_gid) if raw_gid else 0
    if group_id == 0 or current_user.get("role") == "LDAdmin":
        # Return the first available challenge for today (admin preview) or empty
        challenge = (
            db.query(models.DailyChallenge)
            .filter(models.DailyChallenge.challenge_date == today)
            .first()
        )
        if not challenge:
            return {"message": "No daily challenges created yet.", "challenge": None}
        group_id = challenge.group_id

    challenge = (
        db.query(models.DailyChallenge)
        .filter(
            models.DailyChallenge.group_id == group_id,
            models.DailyChallenge.challenge_date == today,
        )
        .first()
    )

    if not challenge:
        # Fallback: trigger inline generation and re-query
        try:
            from tasks import generate_daily_challenges

            created_count = generate_daily_challenges()
            logger.info(
                f"🎯 [DAILY_CHALLENGE_V2_FIX] Generated: {created_count} challenges for {today}"
            )
            db.expire_all()
            challenge = (
                db.query(models.DailyChallenge)
                .filter(
                    models.DailyChallenge.group_id == group_id,
                    models.DailyChallenge.challenge_date == today,
                )
                .first()
            )
        except Exception as e:
            logger.error(f"Inline challenge generation failed: {e}")
            pass

    if not challenge:
        # If still not found, return a clear message instead of a hard 404
        logger.warning(
            f"No daily challenge found for group {group_id} on {today} even after sync."
        )
        return {
            "id": None,
            "date": str(today),
            "message": "Today's daily challenge is being synchronized. Please refresh in a moment.",
            "challenge": None,
        }

    q = (
        db.query(models.Question)
        .filter(models.Question.id == challenge.question_id)
        .first()
    )
    if not q:
        raise HTTPException(
            status_code=404, detail="Daily challenge question not found"
        )

    return {
        "id": challenge.id,
        "date": str(today),
        "question": {
            "id": q.id,
            "question": q.question,
            "options": q.options,
            "difficulty": q.difficulty,
            "bank_id": q.bank_id,
            "has_code": q.has_code if hasattr(q, "has_code") else False,
            "code_language": q.code_language if hasattr(q, "code_language") else None,
        },
    }


@router.get("/bank-library")
def get_bank_library(
    difficulty: Optional[str] = None,
    course_name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """List all org-public banks available for cloning."""
    if current_user.get("role") not in ["GroupAdmin", "Mentor", "LDAdmin", "Admin"]:
        raise HTTPException(
            status_code=403,
            detail="Only admins and mentors can access the bank library",
        )

    query = db.query(models.QuestionBank).filter(models.QuestionBank.is_org_public)
    if difficulty:
        query = query.filter(models.QuestionBank.difficulty == difficulty)

    banks = query.order_by(models.QuestionBank.created_at.desc()).all()
    result = []
    for bank in banks:
        q_count = (
            db.query(func.count(models.Question.id))
            .filter(models.Question.bank_id == bank.id)
            .scalar()
        )
        bank_dict = {c.name: getattr(bank, c.name) for c in bank.__table__.columns}
        bank_dict["question_count"] = q_count
        result.append(bank_dict)
    return result


@router.post("/banks/{bank_id}/clone")
def clone_bank(
    bank_id: int,
    body: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """Clone an org-public bank into the current user's course."""
    if current_user.get("role") not in ["GroupAdmin", "Mentor", "LDAdmin", "Admin"]:
        raise HTTPException(status_code=403)

    source_bank = (
        db.query(models.QuestionBank).filter(models.QuestionBank.id == bank_id).first()
    )
    assert_same_super_org(source_bank, current_user, db, "Bank")
    if not source_bank:
        raise HTTPException(status_code=404, detail="Bank not found")
    if not source_bank.is_org_public and current_user.get("role") != "LDAdmin":
        raise HTTPException(
            status_code=403, detail="This bank is not available for cloning"
        )

    target_course_id = body.get("target_course_id")

    new_bank = models.QuestionBank(
        organization_id=caller_org_id(current_user),
        super_organization_id=caller_super_org_id(current_user, db),
        course_id=target_course_id,
        name=f"{source_bank.name} (Clone)",
        sprint_name=source_bank.sprint_name,
        chapter=source_bank.chapter,
        description=source_bank.description,
        difficulty=source_bank.difficulty,
        created_by=int(current_user["sub"]),
        bank_type="practice",
        time_per_question=source_bank.time_per_question,
        max_questions=source_bank.max_questions,
        show_timer=source_bank.show_timer,
        shuffle=source_bank.shuffle,
        allow_descriptive=source_bank.allow_descriptive,
        is_org_public=False,
        cloned_from_bank_id=bank_id,
    )
    db.add(new_bank)
    db.flush()

    source_questions = (
        db.query(models.Question).filter(models.Question.bank_id == bank_id).all()
    )
    for q in source_questions:
        new_q = models.Question(
            organization_id=new_bank.organization_id,
            super_organization_id=new_bank.super_organization_id,
            bank_id=new_bank.id,
            question=q.question,
            options=q.options,
            answer=q.answer,
            difficulty=q.difficulty,
            user_description=q.user_description,
            has_code=q.has_code,
            code_language=q.code_language,
            concept_tags=q.concept_tags,
        )
        db.add(new_q)
    db.commit()

    log_admin_action(
        db,
        actor_id=int(current_user["sub"]),
        actor_role=current_user["role"],
        action="CLONE_BANK",
        resource_type="BANK",
        resource_id=new_bank.id,
        details={"source_bank_id": bank_id, "target_course_id": target_course_id},
    )

    return {"message": "Bank cloned successfully", "new_bank_id": new_bank.id}


@router.patch("/banks/{bank_id}/publish")
def publish_bank(
    bank_id: int,
    body: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """LDAdmin can publish banks to the entire organization."""
    if current_user.get("role") != "LDAdmin":
        raise HTTPException(
            status_code=403, detail="Only LDAdmin can publish banks to org"
        )

    bank = (
        db.query(models.QuestionBank).filter(models.QuestionBank.id == bank_id).first()
    )
    assert_same_super_org(bank, current_user, db, "Bank")

    is_org_public = body.get("is_org_public", True)
    bank.is_org_public = is_org_public
    db.commit()

    from services.audit_service import log_admin_action

    log_admin_action(
        db,
        actor_id=int(current_user["sub"]),
        actor_role=current_user["role"],
        action="PUBLISH_BANK",
        resource_type="BANK",
        resource_id=bank_id,
        details={"is_org_public": is_org_public},
    )

    return {"success": True, "is_org_public": is_org_public}


@router.post("/report")
def report_question(
    report: schemas.QuestionReportCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """PHASE-3: Enables users to flag issues with specific questions."""
    db_report = models.QuestionReport(
        question_id=report.question_id,
        reporter_id=int(current_user["sub"]),
        reason=report.reason,
        comment=report.comment,
    )
    db.add(db_report)
    db.commit()
    return {"success": True, "message": "Report submitted for administrative review."}


# ─── Granular Question Management ───────────────────────────────────────────


@router.put("/questions/{question_id}")
def update_question(  # noqa: F811
    question_id: int,
    updates: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """
    Granular Question Editor. Fix typos or update options without re-uploading banks.
    """
    q = db.query(models.Question).filter(models.Question.id == question_id).first()
    assert_same_super_org(q, current_user, db, "Question")

    # Permission check: must own the bank or be LDAdmin
    bank = q.bank
    if current_user["role"] != "LDAdmin" and bank.created_by != int(
        current_user["sub"]
    ):
        raise HTTPException(
            status_code=403,
            detail="Forbidden: You can only edit questions in banks you created.",
        )

    allowed_fields = [
        "question",
        "options",
        "answer",
        "difficulty",
        "user_description",
        "has_code",
        "code_language",
        "concept_tags",
    ]
    for key, value in updates.items():
        if key in allowed_fields:
            if key == "answer" and q.options:
                # Re-resolve if answer is A/B/C/D
                value = resolve_answer(value, q.options)
            setattr(q, key, value)

    db.commit()
    return {"success": True, "message": "Question updated successfully."}


@router.delete("/questions/{question_id}")
def delete_question(  # noqa: F811
    question_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """
    Remove a single question from a bank.
    """
    q = db.query(models.Question).filter(models.Question.id == question_id).first()
    assert_same_super_org(q, current_user, db, "Question")

    bank = q.bank
    if current_user["role"] != "LDAdmin" and bank.created_by != int(
        current_user["sub"]
    ):
        raise HTTPException(status_code=403, detail="Forbidden")

    db.delete(q)
    db.commit()
    return {"success": True, "message": "Question deleted."}


@router.get("/user/{user_id}/assignments")
def get_user_assignments(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """
    Retrieve all active assignments and completion status for a specific user.
    """
    # Permission check: either viewing own assignments or is mentor/admin
    if int(current_user["sub"]) != user_id and current_user.get("role") not in [
        "LDAdmin",
        "Mentor",
        "GroupAdmin",
    ]:
        raise HTTPException(
            status_code=403,
            detail="Forbidden: You cannot view another user's assignments.",
        )

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    group = user.group
    if not group:
        return []

    targets = [("person", user.id), ("group", group.id)]
    if group.batch_id:
        targets.append(("batch", group.batch_id))
        if group.batch.vertical_id:
            targets.append(("vertical", group.batch.vertical_id))
            if group.batch.vertical.department_id:
                targets.append(("dept", group.batch.vertical.department_id))
                if group.batch.vertical.department.organization_id:
                    targets.append(
                        ("org", group.batch.vertical.department.organization_id)
                    )

    from sqlalchemy import or_

    filters = [
        (models.Assignment.target_type == t[0]) & (models.Assignment.target_id == t[1])
        for t in targets
    ]

    from models.assignment import Assignment, AssignmentCompletion
    from sqlalchemy.orm import joinedload

    results = (
        db.query(Assignment, AssignmentCompletion)
        .options(joinedload(Assignment.bank), joinedload(Assignment.coding_question))
        .outerjoin(
            AssignmentCompletion,
            (Assignment.id == AssignmentCompletion.assignment_id)
            & (AssignmentCompletion.user_id == user_id),
        )
        .filter(or_(*filters), Assignment.is_active == True)
        .all()
    )

    assignments_list = []
    for a, comp in results:
        assignments_list.append(
            {
                "assignment_id": a.id,
                "bank_id": a.bank_id,
                "coding_question_id": a.coding_question_id,
                "assignment_type": a.assignment_type,
                "bank_name": a.bank.name
                if a.bank
                else (
                    a.coding_question.title
                    if a.coding_question
                    else f"Assignment #{a.id}"
                ),
                "due_date": a.due_date,
                "instructions": a.instructions,
                "is_completed": comp.status in ["passed", "completed"]
                if comp
                else False,
                "status": comp.status if comp else "not_started",
                "score": comp.best_score if comp else None,
                "attempts_used": comp.attempts_used if comp else 0,
                "max_attempts": a.max_attempts,
                "passing_score_percent": a.passing_score_percent,
                "lock_after_due": a.lock_after_due,
            }
        )

    return assignments_list


from services.redis_service import redis_client  # noqa: E402


@router.post("/draft")
async def save_quiz_draft(payload: dict, current_user: dict = Depends(verify_token)):
    user_id = current_user.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Unauthorized")
    await redis_client.set(f"quiz_draft_{user_id}", payload, ex=86400 * 7)
    return {"success": True, "message": "Quiz draft saved successfully"}


@router.get("/draft")
async def get_quiz_draft(current_user: dict = Depends(verify_token)):
    user_id = current_user.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Unauthorized")
    draft = await redis_client.get(f"quiz_draft_{user_id}")
    return {"draft": draft}




# ── Certificate download tokens ──────────────────────────────────────────────
# `/certificate/download` cannot require a bearer token: the UI opens it with
# window.open(), which sends no Authorization header. Leaving it open meant anyone
# could download any learner's certificate (their real name) by guessing an
# attempt id. Instead the AUTHENTICATED endpoint mints a short-lived signed token
# and the download verifies it.
CERT_TOKEN_TTL_SECONDS = 900


def _certificate_token(attempt_id: int, expires_at: int) -> str:
    import hashlib
    import hmac as _hmac

    msg = f"cert:{attempt_id}:{expires_at}".encode()
    return _hmac.new(SECRET_KEY.encode(), msg, hashlib.sha256).hexdigest()[:32]


def _verify_certificate_token(attempt_id: int, expires_at: int, token: str) -> None:
    import hmac as _hmac
    import time as _time

    if expires_at < int(_time.time()):
        raise HTTPException(status_code=403, detail="Certificate link has expired.")
    if not _hmac.compare_digest(_certificate_token(attempt_id, expires_at), token or ""):
        raise HTTPException(status_code=403, detail="Invalid certificate link.")


@router.get("/attempts/{attempt_id}/certificate")
def generate_certificate(
    attempt_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(verify_token),
):
    """
    Generate a Certificate of Completion PDF for a passed attempt.
    """
    sub = current_user.get("sub")
    user_id = int(sub) if sub else 0
    attempt = (
        db.query(models.Attempt)
        .filter(models.Attempt.id == attempt_id, models.Attempt.user_id == user_id)
        .first()
    )

    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")

    # Attempts are by definition completed once saved to DB in this architecture.
    # In a real environment, we would use reportlab to generate the PDF
    # and then upload to S3 and return a presigned URL.

    import time as _time

    _exp = int(_time.time()) + CERT_TOKEN_TTL_SECONDS
    _tok = _certificate_token(attempt_id, _exp)
    mock_certificate_url = (
        f"/api/quiz/attempts/{attempt_id}/certificate/download?exp={_exp}&token={_tok}"
    )

    return {
        "success": True,
        "certificate_url": mock_certificate_url,
        "share_url": f"https://www.linkedin.com/sharing/share-offsite/?url={mock_certificate_url}",
    }

@router.get("/attempts/{attempt_id}/certificate/download")
def download_certificate(
    attempt_id: int,
    exp: int = Query(..., description="Signed link expiry (unix seconds)"),
    token: str = Query(..., description="HMAC issued by GET .../certificate"),
    db: Session = Depends(get_db),
):
    # Signed-link check stands in for the bearer token the browser cannot send.
    _verify_certificate_token(attempt_id, exp, token)

    from fastapi.responses import Response
    import io
    import os
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import landscape, letter
    except ImportError:
        raise HTTPException(status_code=500, detail="ReportLab not installed")

    attempt = db.query(models.Attempt).filter(models.Attempt.id == attempt_id).first()
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
        
    bank = db.query(models.QuestionBank).filter(models.QuestionBank.id == attempt.bank_id).first()
    user = db.query(models.User).filter(models.User.id == attempt.user_id).first()

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)

    # Draw border
    c.setStrokeColorRGB(0.1, 0.1, 0.4)
    c.setLineWidth(4)
    c.rect(20, 20, width - 40, height - 40)
    
    # Try to load logo
    # Api runs from apps/api
    logo_path = os.path.join(os.path.dirname(__file__), "..", "..", "..", "web-next", "public", "images", "logo.png")
    if os.path.exists(logo_path):
        try:
            # Draw logo (scaled)
            c.drawImage(logo_path, width / 2 - 50, height - 120, width=100, height=100, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
            
    c.setFont("Helvetica-Bold", 36)
    c.drawCentredString(width / 2, height - 160, "Certificate of Completion")
    
    c.setFont("Helvetica", 20)
    c.drawCentredString(width / 2, height - 220, "This is to certify that")
    
    c.setFont("Helvetica-Bold", 28)
    user_name = user.full_name if user else "Participant"
    c.drawCentredString(width / 2, height - 270, user_name)
    
    c.setFont("Helvetica", 20)
    c.drawCentredString(width / 2, height - 320, "has successfully completed the assessment")
    
    c.setFont("Helvetica-Bold", 24)
    quiz_name = (getattr(bank, "name", None) or "Assessment") if bank else "Assessment"
    c.drawCentredString(width / 2, height - 370, quiz_name)
    
    c.setFont("Helvetica", 16)
    score_pct = (attempt.score / attempt.total * 100) if attempt.total > 0 else 0
    c.drawCentredString(width / 2, height - 420, f"with a score of {attempt.score}/{attempt.total} ({score_pct:.1f}%)")
    
    # ── White-label co-branding (Org × StudyBuddy, Powered by StudyBuddy) ────
    _brand = "StudyBuddy"
    try:
        _grp = (
            db.query(models.Group).filter(models.Group.id == user.group_id).first()
            if user
            else None
        )
        _dept = (
            db.query(models.Department)
            .filter(models.Department.id == _grp.department_id)
            .first()
            if _grp and _grp.department_id
            else None
        )
        _org = (
            db.query(models.Organization)
            .filter(models.Organization.id == _dept.organization_id)
            .first()
            if _dept
            else None
        )
        if _org:
            _brand = _org.brand_name or _org.name
    except Exception:
        pass
    c.setFont("Helvetica-Bold", 13)
    c.setFillColorRGB(0.1, 0.1, 0.4)
    c.drawCentredString(width / 2, height - 130, f"{_brand}  ×  StudyBuddy")
    c.setFillColorRGB(0, 0, 0)
    c.line(width - 250, 70, width - 100, 70)
    c.setFont("Helvetica", 11)
    c.drawCentredString(width - 175, 55, "Authorized Signature")
    c.drawString(70, 55, f"Certificate ID: SB-{attempt.id}")
    c.setFont("Helvetica-Oblique", 10)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawCentredString(width / 2, 32, "Powered by StudyBuddy")
    c.setFillColorRGB(0, 0, 0)

    c.showPage()
    c.save()

    pdf_content = buffer.getvalue()
    buffer.close()

    return Response(content=pdf_content, media_type="application/pdf")

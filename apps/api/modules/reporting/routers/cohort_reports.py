"""cohort_reports endpoints (moved verbatim from routers/reports.py)."""
from fastapi import APIRouter

from modules.reporting.routers.reports_shared import *  # noqa: F401,F403

router = APIRouter()

@router.get("/batch/{batch_id}/summary")
async def get_batch_report(
    batch_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    assert_batch_in_org(batch_id, db, current_user)
    batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == batch_id).first())
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    groups = await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.batch_id == batch_id).all())
    group_ids = [g.id for g in groups]

    # Date filtering logic. The query is built conditionally, so it is assembled
    # AND executed inside one run_sync call — a lambda cannot hold the branches.
    def _load_attempts(sync_db):
        q = (
            sync_db.query(models.Attempt, models.User)
            .join(models.User, models.Attempt.user_id == models.User.id)
            .filter(models.User.group_id.in_(group_ids))
        )
        if start_date:
            q = q.filter(
                models.Attempt.attempted_at
                >= datetime.datetime.fromisoformat(start_date)
            )
        if end_date:
            q = q.filter(
                models.Attempt.attempted_at
                <= datetime.datetime.fromisoformat(end_date)
            )
        return q.all()

    # STRAT-FIX: Redis Caching for Batch Reports (Section 5.6)
    redis_key = f"batch_report:{batch_id}:{start_date}:{end_date}"
    try:
        cached = await redis_client.get(redis_key)
        if cached:
            return json.loads(cached)
    except Exception as e:
        logger.warning(f"Redis cache lookup failed for batch_report: {e}")
        pass

    attempts = await db.run_sync(_load_attempts)
    total_attempts = len(attempts)
    total_members = (
        await db.run_sync(lambda s: s.query(models.User).filter(models.User.group_id.in_(group_ids)).count())
    )

    avg_accuracy = 0.0
    if total_attempts > 0:
        total_score = sum((a.score or 0) for a, _ in attempts)
        total_points = sum((a.total or 0) for a, _ in attempts)
        avg_accuracy = (
            round((total_score / total_points * 100), 2) if total_points > 0 else 0.0
        )

    # Build per-user stats for leaderboard with attempt count
    user_stats: dict = {}
    for a, u in attempts:
        uid = u.id
        if uid not in user_stats:
            user_stats[uid] = {
                "full_name": u.full_name,
                "group_name": u.group.name if u.group else "Member",
                "scores": [],
                "attempt_count": 0,
            }
        acc = (a.score / a.total * 100) if (a.total and a.total > 0) else 0
        user_stats[uid]["scores"].append(acc)
        user_stats[uid]["attempt_count"] += 1

    top_performers = sorted(
        [
            {
                "full_name": s["full_name"],
                "group_name": s["group_name"],
                "avg_score": round(sum(s["scores"]) / len(s["scores"]), 1)
                if s["scores"]
                else 0,
                "attempt_count": s["attempt_count"],
            }
            for s in user_stats.values()
        ],
        key=lambda x: x["avg_score"],
        reverse=True,
    )[:5]

    # Per-group breakdown — avoid N+1 by building a map first
    group_attempt_map: dict = {g.id: [] for g in groups}
    for a, u in attempts:
        if u.group_id in group_attempt_map:
            group_attempt_map[u.group_id].append(a)

    group_breakdown = []
    for g in groups:
        g_atts = group_attempt_map[g.id]
        g_total = len(g_atts)
        g_avg_acc = 0.0
        if g_total > 0:
            g_score = sum((a.score or 0) for a in g_atts)
            g_points = sum((a.total or 0) for a in g_atts)
            g_avg_acc = round((g_score / g_points * 100), 2) if g_points > 0 else 0.0
        group_breakdown.append(
            {
                "id": g.id,
                "group_name": g.name,
                "attempts": g_total,
                "avg_score": g_avg_acc,
            }
        )

    # Generate Strategic Observations using AI
    stats_for_ai = {
        "average_score": avg_accuracy,
        "total_members": total_members,
        "total_attempts": total_attempts,
        "group_performance": group_breakdown,
        "top_performers": top_performers,
    }
    vertical_name = await db.run_sync(
        lambda sync_db: (
            sync_db.query(models.Vertical.name)
            .filter(models.Vertical.id == batch.vertical_id)
            .scalar()
        )
        or "N/A"
    )

    observations = await ai_executive.generate_batch_insights(batch.name, stats_for_ai)

    report = {
        "batch_name": batch.name,
        # `batch.vertical` is a lazy relationship; an AsyncSession cannot resolve
        # it implicitly (MissingGreenlet). Resolved explicitly above.
        "vertical_name": vertical_name,
        "total_members": total_members,
        "total_groups": len(groups),
        "total_attempts": total_attempts,
        "average_score": avg_accuracy,
        "top_performers": top_performers,
        "group_performance": group_breakdown,
        "strategic_observations": observations,
        "from_cache": False,
    }

    try:
        await redis_client.set(redis_key, json.dumps(report), ex=21600)  # 6h TTL
    except Exception as e:
        logger.warning(f"Redis cache lookup failed for group_report: {e}")
        pass

    return report

@router.get("/group/{group_id}/cohort-health")
async def get_cohort_health(
    group_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(verify_token),
):
    """PHASE-3: Generate 10 targeted strategic intervention points for a group."""
    assert_group_in_org(group_id, db, current_user)
    group = await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.id == group_id).first())
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    # Authorization: Admins/Mentors assigned to this group
    if current_user["role"] not in ["LDAdmin", "Mentor", "GroupAdmin"]:
        raise HTTPException(status_code=403)

    # Build performance context
    users = (
        await db.run_sync(lambda s: s.query(models.User)
        .filter(models.User.group_id == group_id, models.User.is_active.is_(True))
        .all())
    )
    user_ids = [u.id for u in users]

    attempts = (
        await db.run_sync(lambda s: s.query(models.Attempt).filter(models.Attempt.user_id.in_(user_ids)).all())
    )
    # FIX #11: zero guard on sum(a.total) to prevent ZeroDivisionError
    _total_sum = sum(a.total for a in attempts if a.total)
    avg_accuracy = (
        (sum(a.score for a in attempts) / _total_sum * 100) if _total_sum > 0 else 0
    )

    metrics = {
        "avg_accuracy": round(avg_accuracy, 1),
        "active_members": len(users),
        "total_attempts": len(attempts),
        "at_risk_count": len([u for u in users if getattr(u, "streak_count", 0) == 0]),
        "top_performer": max([u.full_name for u in users], default="N/A"),
    }

    observations = await ai_executive.generate_cohort_health(group.name, metrics)
    return {
        "group_name": group.name,
        "metrics": metrics,
        "strategic_interventions": observations,
    }

@router.post("/group/{group_id}/refresh-intelligence")
async def refresh_group_intelligence(
    group_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    """PHASE-3: Force re-calculate and re-cache all performance vectors for a group."""
    assert_group_in_org(group_id, db, current_user)
    from cache_manager import cache_manager

    users = await db.run_sync(lambda s: s.query(models.User).filter(models.User.group_id == group_id).all())

    count = 0
    from services.performance_engine import performance_engine

    for u in users:
        # Clear specific user cache
        await cache_manager.invalidate(f"user_vectors:{u.id}")
        # Re-warm cache
        await performance_engine.get_user_vectors(u.id, db, refresh=True)
        count += 1

    return {
        "success": True,
        "refreshed_count": count,
        "message": f"Intelligence re-synchronized for {count} members.",
    }

@router.get("/batch/{batch_id}/xlsx")
async def export_batch_xlsx(
    batch_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    """Professional Multi-Sheet L&D Executive Export."""
    assert_batch_in_org(batch_id, db, current_user)
    from cache_manager import redis_client

    lock_key = f"rl:export_batch_xlsx:{current_user['sub']}"
    try:
        acquired = await redis_client.set(lock_key, "locked", ex=30)
        if not acquired:
            raise HTTPException(
                status_code=429,
                detail="Export already in progress or requested too recently. Please wait.",
            )
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        pass

    import io

    from fastapi.responses import StreamingResponse

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == batch_id).first())
    if not batch:
        raise HTTPException(status_code=404)

    groups = await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.batch_id == batch_id).all())
    group_ids = [g.id for g in groups]

    # Fetch data and generate AI insights first
    report_data = await get_batch_report(
        batch_id=batch_id, db=db, current_user=current_user
    )
    rows = (
        await db.run_sync(lambda s: s.query(models.Attempt, models.User)
        .join(models.User)
        .filter(models.User.group_id.in_(group_ids))
        .all())
    )

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")
    Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Sheet 1: Executive Summary ---
    ws_summary = wb.active
    assert ws_summary is not None
    ws_summary.title = "Executive Summary"

    ws_summary.merge_cells("A1:E1")
    ws_summary["A1"] = f"STRATEGIC REPORT: {batch.name.upper()}"
    ws_summary["A1"].font = Font(bold=True, size=16, color="3730A3")
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    ws_summary.append([])
    ws_summary.append(["METRIC", "CURRENT STATUS"])
    for cell in ws_summary[3]:
        cell.font = header_font
        cell.fill = header_fill

    ws_summary.append(["Global Proficiency", f"{report_data['average_score']}%"])
    ws_summary.append(["Total Active Members", report_data["total_members"]])
    ws_summary.append(["Engagement Index (Attempts)", report_data["total_attempts"]])
    ws_summary.append(
        [
            "Report Iteration",
            datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M"),
        ]
    )

    ws_summary.append([])
    ws_summary.append(["AI STRATEGIC OBSERVATIONS"])
    ws_summary.merge_cells(f"A{ws_summary.max_row}:E{ws_summary.max_row}")
    ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True, color="FFFFFF")
    ws_summary[f"A{ws_summary.max_row}"].fill = PatternFill("solid", fgColor="4338CA")

    for obs in report_data.get("strategic_observations", []):
        ws_summary.append([f"• {obs}"])
        ws_summary.merge_cells(f"A{ws_summary.max_row}:E{ws_summary.max_row}")

    # --- Sheet 2: Group Performance ---
    ws_groups = wb.create_sheet("Group Performance")
    ws_groups.append(
        ["GROUP NAME", "ATTEMPTS", "AVG PROFICIENCY %", "ENGAGEMENT LEVEL"]
    )
    for cell in ws_groups[1]:
        cell.font = header_font
        cell.fill = header_fill

    for g in report_data["group_performance"]:
        status = (
            "EXCEPTIONAL"
            if g["avg_score"] > 85
            else "OPTIMAL"
            if g["avg_score"] > 70
            else "NEEDS SYNC"
        )
        ws_groups.append([g["group_name"], g["attempts"], f"{g['avg_score']}%", status])

    # --- Sheet 3: Individual Leaderboard ---
    ws_members = wb.create_sheet("Member Registry")
    ws_members.append(
        ["MEMBER NAME", "GROUP", "TOTAL ATTEMPTS", "AVG ACCURACY %", "TOP SCORE"]
    )
    for cell in ws_members[1]:
        cell.font = header_font
        cell.fill = header_fill

    for p in report_data["top_performers"]:
        ws_members.append(
            [
                p["full_name"],
                p["group_name"],
                p["attempt_count"],
                f"{p['avg_score']}%",
                "-",
            ]
        )

    # --- Sheet 4: Raw Logs ---
    ws_raw = wb.create_sheet("Raw Activity Logs")
    headers = [
        "Attempt ID",
        "User ID",
        "Full Name",
        "Group",
        "Score",
        "Total",
        "Accuracy %",
        "Timestamp",
    ]
    ws_raw.append(headers)
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill

    for a, u in rows:
        acc = round((a.score / a.total * 100), 1) if a.total > 0 else 0
        ws_raw.append(
            [
                a.id,
                u.id,
                u.full_name,
                u.group.name if u.group else "N/A",
                a.score,
                a.total,
                acc,
                a.attempted_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    from openpyxl.utils import get_column_letter

    for sheet in wb.worksheets:
        for col in sheet.columns:
            if not col:
                continue
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # type: ignore
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    logger.warning(f"Column width calculation failed: {e}")
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=batch_{batch_id}_executive_report.xlsx"
        },
    )

@router.get("/compare")
def compare_batches(
    batch_ids: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_ldadmin),
):
    """PHASE-3: Side-by-side benchmarking for multiple batches."""
    ids = [int(i.strip()) for i in batch_ids.split(",") if i.strip().isdigit()]
    result = []
    for bid in ids:
        batch = db.query(models.Batch).filter(models.Batch.id == bid).first()
        if not batch:
            continue

        g_ids = [
            g.id
            for g in db.query(models.Group).filter(models.Group.batch_id == bid).all()
        ]

        # Quiz Stats
        attempts = (
            db.query(models.Attempt)
            .join(models.User)
            .filter(models.User.group_id.in_(g_ids))
            .all()
        )
        q_acc = (
            sum(a.score for a in attempts) / sum(a.total for a in attempts) * 100
            if attempts and sum(a.total for a in attempts) > 0
            else 0
        )

        # Coding Stats
        c_atts = (
            db.query(models.CodingAttempt)
            .join(models.User)
            .filter(models.User.group_id.in_(g_ids))
            .all()
        )
        c_acc = (
            sum(ca.score for ca in c_atts if ca.score)
            / len([ca for ca in c_atts if ca.score])
            if c_atts and len([ca for ca in c_atts if ca.score]) > 0
            else 0
        )

        result.append(
            {
                "batch_id": bid,
                "batch_name": batch.name,
                "quiz_accuracy": round(q_acc, 1),
                "coding_proficiency": round(c_acc, 1),
                "active_members": db.query(models.User)
                .filter(models.User.group_id.in_(g_ids))
                .count(),
                "engagement_score": round(
                    (len(attempts) + len(c_atts)) / max(1, len(g_ids)), 1
                ),
            }
        )
    return result

@router.get("/group/{group_id}/health")
def get_group_health(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_mentor_or_above),
):
    # Section 2: Scope check for Mentors (SEC-102)
    assert_group_in_org(group_id, db, current_user)
    if current_user.get("role") == "Mentor":
        user_id = int(current_user["sub"])
        # Check V3 UserRole table (Strategic Mapping)
        exists = (
            db.query(models.UserRole)
            .filter(
                models.UserRole.user_id == user_id,
                models.UserRole.role == "Mentor",
                models.UserRole.scope_type == "group",
                models.UserRole.scope_id == group_id,
            )
            .first()
        )

        if not exists:
            # Fallback to V2 legacy assignment table (Cross-version Compatibility)
            assign = (
                db.query(models.MentorGroupAssignment)
                .filter_by(mentor_id=user_id, group_id=group_id, is_active=True)
                .first()
            )
            if not assign:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Access denied: You do not have Mentor oversight for this sector.",
                )

    group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    # Stats for the group
    attempts = (
        db.query(models.Attempt)
        .join(models.User)
        .filter(models.User.group_id == group_id)
        .all()
    )

    # Chapter-wise breakdown
    chapters = {}
    for a in attempts:
        bank = (
            db.query(models.QuestionBank)
            .filter(models.QuestionBank.id == a.bank_id)
            .first()
        )
        if bank and bank.chapter:
            if bank.chapter not in chapters:
                chapters[bank.chapter] = {"total_q": 0, "correct_q": 0}
            chapters[bank.chapter]["total_q"] += a.total
            chapters[bank.chapter]["correct_q"] += a.score

    health_data = []
    for ch, stats in chapters.items():
        acc = (
            (stats["correct_q"] / stats["total_q"]) * 100 if stats["total_q"] > 0 else 0
        )
        health_data.append({"chapter": ch, "accuracy": round(acc, 2)})

    return {
        "group_id": group_id,
        "group_name": group.name,
        "health": health_data,
        "participation_rate": round((len(attempts) / (len(group.users) or 1)) * 100, 2)
        if group.users
        else 0,
    }

@router.get("/analytics/comparative")
async def get_comparative_analytics(
    db: AsyncSession = Depends(get_async_db), current_user: dict = Depends(require_ldadmin)
):
    """Global KPIs for the L&D Admin dashboard."""
    import json

    from cache_manager import redis_client

    redis_key = "reports:comparative_analytics"
    try:
        cached = await redis_client.get(redis_key)
        if cached:
            return json.loads(cached)
    except Exception:
        pass
    total_users = await db.run_sync(lambda s: s.query(models.User).count())
    active_users_30d = (
        await db.run_sync(lambda s: s.query(models.User)
        .join(models.Attempt)
        .filter(
            models.Attempt.attempted_at
            >= datetime.datetime.now(datetime.timezone.utc)
            - datetime.timedelta(days=30)
        )
        .distinct()
        .count())
    )

    # Unified attempts (PHASE-3 alignment)
    attempts = await db.run_sync(lambda s: s.query(models.Attempt).all())
    coding_attempts = await db.run_sync(lambda s: s.query(models.CodingAttempt).all())

    total_score = sum((a.score or 0) for a in attempts)
    total_points = sum((a.total or 0) for a in attempts)

    # Add coding scores (scaled to 100 for consistency)
    total_score += sum(
        (ca.score or 0) for ca in coding_attempts if ca.score is not None
    )
    total_points += sum(100 for ca in coding_attempts if ca.score is not None)

    # Calculate weak topics (accuracy < 60%)
    banks = await db.run_sync(lambda s: s.query(models.QuestionBank).all())
    weak_count = 0
    topic_accuracy = {}
    for bank in banks:
        bank_atts = [a for a in attempts if a.bank_id == bank.id]
        if not bank_atts:
            continue
        b_total_score = sum((a.score or 0) for a in bank_atts)
        b_total_points = sum((a.total or 0) for a in bank_atts)
        bank_acc = (b_total_score / b_total_points) * 100 if b_total_points > 0 else 0
        topic_accuracy[bank.name] = bank_acc
        if bank_acc < 60:
            weak_count += 1

    # Add coding topics to trends
    coding_questions = await db.run_sync(lambda s: s.query(models.CodingQuestion).all())
    for cq in coding_questions:
        cq_atts = [
            ca
            for ca in coding_attempts
            if getattr(ca, "coding_question_id") == getattr(cq, "id") and getattr(ca, "score") is not None
        ]
        if not cq_atts:
            continue
        cq_acc = sum(getattr(ca, "score") or 0 for ca in cq_atts) / len(cq_atts)
        topic_accuracy[f"Code: {cq.title}"] = cq_acc
        if cq_acc < 60:
            weak_count += 1

    # --- Enhanced Multi-Level Performance ---
    # Eager-load departments: `org.departments` is traversed below, and a lazy load
    # outside run_sync raises MissingGreenlet on an AsyncSession.
    orgs = await db.run_sync(
        lambda s: s.query(models.Organization)
        .options(selectinload(models.Organization.departments))
        .all()
    )
    org_stats = []
    for org in orgs:
        # Get all users in this org
        dept_ids = [d.id for d in org.departments]
        vert_ids = [
            v.id
            for v in await db.run_sync(lambda s: s.query(models.Vertical)
            .filter(models.Vertical.department_id.in_(dept_ids))
            .all())
        ]
        batch_ids = [
            b.id
            for b in await db.run_sync(lambda s: s.query(models.Batch)
            .filter(models.Batch.vertical_id.in_(vert_ids))
            .all())
        ]
        group_ids = [
            g.id
            for g in await db.run_sync(lambda s: s.query(models.Group)
            .filter(models.Group.batch_id.in_(batch_ids))
            .all())
        ]

        org_attempts = (
            await db.run_sync(lambda s: s.query(models.Attempt)
            .join(models.User)
            .filter(models.User.group_id.in_(group_ids))
            .all())
        )
        if not org_attempts:
            continue

        o_score = sum((a.score or 0) for a in org_attempts)
        o_points = sum((a.total or 0) for a in org_attempts)
        o_acc = (o_score / o_points * 100) if o_points > 0 else 0
        org_stats.append({"label": org.name, "value": round(o_acc, 1)})

    # Trends: most active banks/topics
    sorted_topics = sorted(topic_accuracy.items(), key=lambda x: x[1], reverse=True)[:5]
    recent_trends = [
        {
            "label": t[0],
            "value": round(t[1], 1),
            "color": "text-emerald-400"
            if t[1] > 80
            else "text-indigo-400"
            if t[1] > 60
            else "text-rose-400",
        }
        for t in sorted_topics
    ]

    # Calculate real uptake trend (last 30d vs 30d-60d)
    now = datetime.datetime.now(datetime.timezone.utc)
    # Combine user activity from both types. The sub-queries are Query objects
    # combined with .union(), so they must be built and consumed together inside
    # a single run_sync block.
    def _active_counts(sync_db):
        q30 = sync_db.query(models.Attempt.user_id).filter(
            models.Attempt.attempted_at >= now - datetime.timedelta(days=30)
        )
        c30 = sync_db.query(models.CodingAttempt.user_id).filter(
            models.CodingAttempt.attempted_at >= now - datetime.timedelta(days=30)
        )
        last_30 = (
            sync_db.query(models.User)
            .filter(models.User.id.in_(q30.union(c30)))
            .count()
        )

        q60 = sync_db.query(models.Attempt.user_id).filter(
            models.Attempt.attempted_at < now - datetime.timedelta(days=30),
            models.Attempt.attempted_at >= now - datetime.timedelta(days=60),
        )
        c60 = sync_db.query(models.CodingAttempt.user_id).filter(
            models.CodingAttempt.attempted_at < now - datetime.timedelta(days=30),
            models.CodingAttempt.attempted_at >= now - datetime.timedelta(days=60),
        )
        prev_30 = (
            sync_db.query(models.User)
            .filter(models.User.id.in_(q60.union(c60)))
            .count()
        )
        return last_30, prev_30

    active_30d, active_60d_30d = await db.run_sync(_active_counts)

    diff = active_30d - active_60d_30d
    trend_symbol = "+" if diff >= 0 else "-"
    uptake_trend = f"{trend_symbol}{abs(diff)} Active Delta"

    res = {
        "system_uptake": round((active_users_30d / total_users * 100), 1)
        if total_users > 0
        else 0,
        "uptake_trend": uptake_trend,
        "global_avg_accuracy": round((total_score / total_points * 100), 1)
        if total_points > 0
        else 0,
        "average_proficiency": round((total_score / total_points * 100), 1)
        if total_points > 0
        else 0,
        "active_users": active_users_30d,
        "weak_topics_count": weak_count,
        "recent_trends": recent_trends,
        "org_performance": org_stats[:5],
        "health_status": "Cluster Operational"
        if active_users_30d > 0
        else "Low Activity",
    }

    try:
        await redis_client.set(redis_key, json.dumps(res), ex=3600)
    except Exception:
        pass

    return res

@router.get("/lnd/stats")
async def get_lnd_stats(
    db: AsyncSession = Depends(get_async_db), current_user: dict = Depends(require_ldadmin)
):
    """Explicit alias for the L&D Admin summary stats."""
    return await get_comparative_analytics(db, current_user)

@router.get("/analytics/performance-distribution")
async def get_performance_distribution(
    batch_id: int | None = None,
    group_id: int | None = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    """
    Section 12 Method #11 — Performance Distribution (Z-Score Analysis)
    Z = (x - μ) / σ per user
    """
    assert_batch_in_org(batch_id, db, current_user)
    assert_group_in_org(group_id, db, current_user)
    import json

    redis_key = f"reports:perf_dist:{batch_id}:{group_id}"
    try:
        cached = await redis_client.get(redis_key)
        if cached:
            return json.loads(cached)
    except Exception:
        pass

    # Conditional query assembly + execution in one run_sync block.
    def _load(sync_db):
        q = sync_db.query(models.Attempt, models.User).join(models.User)
        if group_id:
            q = q.filter(models.User.group_id == group_id)
        elif batch_id:
            group_ids = [
                g.id
                for g in sync_db.query(models.Group)
                .filter(models.Group.batch_id == batch_id)
                .all()
            ]
            q = q.filter(models.User.group_id.in_(group_ids))
        return q.all()

    attempts = await db.run_sync(_load)
    if not attempts:
        return {"distribution": [], "mean": 0, "std_dev": 0}

    user_scores = {}
    for a, u in attempts:
        acc = (a.score / a.total * 100) if a.total and a.total > 0 else 0
        uid = u.id
        if uid not in user_scores:
            user_scores[uid] = {"full_name": u.full_name, "scores": []}
        user_scores[uid]["scores"].append(acc)

    user_avgs = {
        uid: sum(v["scores"]) / len(v["scores"]) for uid, v in user_scores.items()
    }
    all_scores = list(user_avgs.values())
    mean = sum(all_scores) / len(all_scores) if all_scores else 0
    variance = (
        sum((s - mean) ** 2 for s in all_scores) / len(all_scores) if all_scores else 0
    )
    std_dev = math.sqrt(variance)

    distribution = []
    for uid, avg in user_avgs.items():
        z = (avg - mean) / std_dev if std_dev > 0 else 0
        quadrant = (
            "Star"
            if avg >= mean and z > 0.5
            else "Solid Performer"
            if avg >= mean
            else "Rising Star"
            if z > -0.5
            else "At-Risk"
        )
        distribution.append(
            {
                "user_id": uid,
                "full_name": user_scores[uid]["full_name"],
                "avg_score": round(avg, 1),
                "z_score": round(z, 2),
                "quadrant": quadrant,
            }
        )

    res = {
        "distribution": sorted(
            distribution, key=lambda x: x["avg_score"], reverse=True
        ),
        "mean": round(mean, 1),
        "std_dev": round(std_dev, 2),
        "cohort_size": len(distribution),
    }
    try:
        await redis_client.set(redis_key, json.dumps(res), ex=3600)
    except Exception:
        pass
    return res

@router.get("/analytics/engagement-decay")
def get_engagement_decay(
    batch_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_ldadmin),
):
    """
    Section 12 Method #4 — Engagement Decay Index (Churn Prediction)
    Rolling 7-day activity vs historical average.
    """
    assert_batch_in_org(batch_id, db, current_user)
    now = datetime.datetime.now(datetime.timezone.utc)
    seven_days_ago = now - datetime.timedelta(days=7)
    thirty_days_ago = now - datetime.timedelta(days=30)

    if batch_id:
        group_ids = [
            g.id
            for g in db.query(models.Group)
            .filter(models.Group.batch_id == batch_id)
            .all()
        ]
        user_filter = models.User.group_id.in_(group_ids)
    else:
        from sqlalchemy import true
        user_filter = true()

    recent_active = (
        db.query(models.User)
        .join(models.Attempt)
        .filter(models.Attempt.attempted_at >= seven_days_ago, user_filter)
        .distinct()
        .count()
    )

    historical_active = (
        db.query(models.User)
        .join(models.Attempt)
        .filter(
            models.Attempt.attempted_at >= thirty_days_ago,
            models.Attempt.attempted_at < seven_days_ago,
            user_filter,
        )
        .distinct()
        .count()
    )

    weekly_avg_historical = (
        historical_active / 3.29 if historical_active > 0 else 0
    )  # ~23 days / 7
    decay_index = (
        ((recent_active - weekly_avg_historical) / weekly_avg_historical * 100)
        if weekly_avg_historical > 0
        else 0
    )

    risk_level = (
        "Low Risk"
        if decay_index >= -10
        else "Medium Risk"
        if decay_index >= -30
        else "High Risk — Churn Likely"
    )

    return {
        "recent_7d_active": recent_active,
        "historical_weekly_avg": round(weekly_avg_historical, 1),
        "decay_index_pct": round(decay_index, 1),
        "risk_level": risk_level,
        "interpretation": f"Engagement is {abs(int(decay_index))}% {'above' if decay_index >= 0 else 'below'} historical baseline.",
    }

@router.get("/analytics/composite-health-index")
def get_composite_health_index(
    batch_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_ldadmin),
):
    """
    Section 12 Method #30 — Composite Learning Health Index (CHI)
    Weighted blend of 8 primary KPIs: accuracy, participation, attempts, weak topics, etc.
    """
    assert_batch_in_org(batch_id, db, current_user)
    total_users = db.query(models.User).count()
    if total_users == 0:
        return {"chi": 0, "components": {}, "grade": "N/A"}

    now = datetime.datetime.now(datetime.timezone.utc)
    active_30d = (
        db.query(models.User)
        .join(models.Attempt)
        .filter(models.Attempt.attempted_at >= now - datetime.timedelta(days=30))
        .distinct()
        .count()
    )

    all_attempts = db.query(models.Attempt).all()
    total_score = sum((a.score or 0) for a in all_attempts)
    total_points = sum((a.total or 0) for a in all_attempts)
    avg_acc = (total_score / total_points * 100) if total_points > 0 else 0

    participation_rate = (active_30d / total_users * 100) if total_users > 0 else 0
    attempt_volume = min(100, len(all_attempts) / max(1, total_users) * 10)  # normalize

    # CHI formula (weighted): accuracy 40%, participation 35%, attempt volume 25%
    chi = round(avg_acc * 0.40 + participation_rate * 0.35 + attempt_volume * 0.25, 1)

    grade = (
        "A"
        if chi >= 85
        else "B"
        if chi >= 70
        else "C"
        if chi >= 55
        else "D"
        if chi >= 40
        else "F"
    )

    return {
        "chi": chi,
        "grade": grade,
        "components": {
            "avg_accuracy_pct": round(avg_acc, 1),
            "participation_rate_pct": round(participation_rate, 1),
            "attempt_volume_score": round(attempt_volume, 1),
        },
        "interpretation": f"Platform Composite Health Index: {chi}/100 (Grade {grade})",
    }

@router.get("/coding-leaderboard")
async def get_coding_leaderboard(
    group_id: int | None = None,
    batch_id: int | None = None,
    page: int = 1,
    page_size: int = 50,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    """
    Coding challenge leaderboard — surfacing CodingAttempt data.
    Previously stored but never exposed to any UI.
    """
    assert_batch_in_org(batch_id, db, current_user)
    assert_group_in_org(group_id, db, current_user)
    import json

    redis_key = f"reports:coding_leaderboard:{group_id}:{batch_id}:{page}:{page_size}"
    try:
        cached = await redis_client.get(redis_key)
        if cached:
            return json.loads(cached)
    except Exception:
        pass

    # Conditional assembly + BOTH terminals (count and page) share one Query, so
    # they are executed together inside a single run_sync block.
    def _load_page(sync_db):
        q = (
            sync_db.query(models.CodingAttempt, models.User, models.CodingQuestion)
            .join(models.User, models.CodingAttempt.user_id == models.User.id)
            .join(
                models.CodingQuestion,
                models.CodingAttempt.coding_question_id == models.CodingQuestion.id,
            )
            .filter(models.CodingAttempt.leaderboard_eligible)
        )
        if group_id:
            q = q.filter(models.User.group_id == group_id)
        elif batch_id:
            gids = [
                g.id
                for g in sync_db.query(models.Group)
                .filter(models.Group.batch_id == batch_id)
                .all()
            ]
            q = q.filter(models.User.group_id.in_(gids))
        return (
            q.count(),
            q.order_by(models.CodingAttempt.score.desc().nullslast())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all(),
        )

    total, rows = await db.run_sync(_load_page)

    res = {
        "total": total,
        "page": page,
        "page_size": page_size,
        "leaderboard": [
            {
                "user_id": u.id,
                "full_name": u.full_name,
                "group_name": u.group.name if u.group else "N/A",
                "question_title": q.title,
                "score": ca.score,
                "criteria_scores": ca.criteria_scores,
                "language": ca.language,
                "submitted_at": ca.attempted_at.isoformat()
                if ca.attempted_at
                else None,
            }
            for ca, u, q in rows
        ],
    }
    try:
        await redis_client.set(redis_key, json.dumps(res), ex=3600)
    except Exception:
        pass
    return res

@router.get("/group-performance-stack")
async def get_group_leaderboard(
    group_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(verify_token),
):
    """
    STRAT-ANALYTICS-01: Comparative leaderboard for group performance.
    Calculates weighted proficiency score for all members in the group.
    """
    assert_group_in_org(group_id, db, current_user)
    # Authorization: Ensure actor has access to this group
    if current_user["role"] != "LDAdmin":
        if int(current_user.get("group_id", -1)) != group_id:
            # Check if actor is a Mentor for this group
            from models.auth import MentorGroupAssignment

            is_assigned = (
                await db.run_sync(lambda s: s.query(MentorGroupAssignment)
                .filter_by(
                    mentor_id=int(current_user["sub"]),
                    group_id=group_id,
                    is_active=True,
                )
                .first())
            )
            if not is_assigned:
                raise HTTPException(
                    status_code=403,
                    detail="Strategic Boundary Violation: You do not have oversight for this sector.",
                )

    import json

    redis_key = f"reports:group_leaderboard:{group_id}"
    try:
        cached = await redis_client.get(redis_key)
        if cached:
            return json.loads(cached)
    except Exception:
        pass

    users = await db.run_sync(
        lambda s: s.query(models.User)
        # `user.group` is traversed in the loop below; a lazy load there raises
        # MissingGreenlet because it happens outside run_sync.
        .options(selectinload(models.User.group))
        .filter(models.User.group_id == group_id, models.User.is_active.is_(True))
        .all()
    )
    leaderboard = []
    import datetime

    now = datetime.datetime.now(datetime.timezone.utc)

    for user in users:
        # Quiz Stats
        quiz_attempts_all = (
            await db.run_sync(lambda s: s.query(models.Attempt).filter(models.Attempt.user_id == user.id).all())
        )
        valid_quiz = [a for a in quiz_attempts_all if a.total and a.total > 0]
        quiz_acc = (
            round(
                sum((a.score / a.total) * 100 for a in valid_quiz) / len(valid_quiz), 1
            )
            if valid_quiz
            else 0
        )

        # Coding Stats
        coding_attempts_all = (
            await db.run_sync(lambda s: s.query(models.CodingAttempt)
            .filter(models.CodingAttempt.user_id == user.id)
            .all())
        )
        code_total = len(coding_attempts_all)
        code_passed = sum(1 for c in coding_attempts_all if c.score and c.score >= 70)
        code_acc = round((code_passed / code_total * 100), 1) if code_total > 0 else 0

        # AI Code Score
        ai_scored = [c for c in coding_attempts_all if getattr(c, "score") is not None]
        avg_ai_score = (
            round(sum(getattr(c, "score") or 0 for c in ai_scored) / len(ai_scored), 1)
            if ai_scored
            else 0
        )

        # Assignment completion (PHASE-3 alignment)
        # Find all assignments targeting this group or its parents (batch/vertical)
        batch_id = user.group.batch_id if user.group else None
        vertical_id = None
        if batch_id:
            batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == batch_id).first())
            vertical_id = batch.vertical_id if batch else None

        target_filters = [
            (models.Assignment.target_type == "group")
            & (models.Assignment.target_id == group_id)
        ]
        if batch_id:
            target_filters.append(
                (models.Assignment.target_type == "batch")
                & (models.Assignment.target_id == batch_id)
            )
        if vertical_id:
            target_filters.append(
                (models.Assignment.target_type == "vertical")
                & (models.Assignment.target_id == vertical_id)
            )

        total_assignments = (
            await db.run_sync(lambda s: s.query(models.Assignment)
            .filter(models.Assignment.is_active.is_(True), or_(*target_filters))
            .count())
        )

        completed_asgn = (
            await db.run_sync(lambda s: s.query(models.AssignmentCompletion)
            .filter(
                models.AssignmentCompletion.user_id == user.id,
            )
            .count())
        )
        asgn_pct = (
            round((completed_asgn / total_assignments) * 100, 1)
            if total_assignments > 0
            else 0
        )

        # Active days
        active_dates = set()
        for a in quiz_attempts_all:
            if a.attempted_at:
                active_dates.add(
                    a.attempted_at.date()
                    if hasattr(a.attempted_at, "date")
                    else a.attempted_at
                )
        for c in coding_attempts_all:
            if getattr(c, "attempted_at"):
                active_dates.add(
                    c.attempted_at.date()
                    if hasattr(c.attempted_at, "date")
                    else c.attempted_at
                )

        # Last active
        last_active = user.last_active_date
        # FIX #12: ensure last_active is timezone-aware before arithmetic with timezone-aware now
        if last_active is not None and last_active.tzinfo is None:
            last_active = last_active.replace(tzinfo=datetime.timezone.utc)
        days_since = (
            int((now - last_active).total_seconds() / 86400) if last_active else None
        )

        # Velocity (simple)
        sorted_attempts = sorted(
            valid_quiz, key=lambda a: a.attempted_at or datetime.datetime.min
        )
        if len(sorted_attempts) >= 4:
            n = len(sorted_attempts) // 2
            v1 = sum((a.score / a.total) * 100 for a in sorted_attempts[:n]) / n
            v2 = sum((a.score / a.total) * 100 for a in sorted_attempts[n:]) / (
                len(sorted_attempts) - n
            )
            velocity = v2 - v1
        else:
            velocity = 0

        # Risk
        if days_since is None or days_since > 14:
            risk = "High Risk"
        elif days_since > 7 or quiz_acc < 40:
            risk = "Medium Risk"
        else:
            risk = "On Track"

        # Weighted Overall (50/30/20 split — quiz/coding/assignment)
        overall = round((quiz_acc * 0.5) + (code_acc * 0.3) + (asgn_pct * 0.2), 1)

        leaderboard.append(
            {
                "user_id": user.id,
                "custom_slug": user.custom_slug,
                "full_name": user.full_name,
                "email": user.email,
                "profile_photo_url": user.profile_photo_url,
                "quiz_accuracy": quiz_acc,
                "coding_accuracy": code_acc,
                "overall_score": overall,
                "ai_avg_score": avg_ai_score,
                "assignment_completion": asgn_pct,
                "streak": user.streak_count or 0,
                "total_quiz_attempts": len(quiz_attempts_all),
                "total_coding_attempts": code_total,
                "days_active": len(active_dates),
                "last_active_days_ago": days_since,
                "velocity": round(velocity, 1),
                "risk_level": risk,
            }
        )

    ranked = sorted(leaderboard, key=lambda x: x["overall_score"], reverse=True)
    for i, entry in enumerate(ranked):
        entry["rank"] = i + 1

    try:
        await redis_client.set(redis_key, json.dumps(ranked), ex=3600)
    except Exception:
        pass

    return ranked

@router.get("/batch/{batch_id}/export")
async def export_batch_report(
    batch_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    """FUNC-006: Professional multi-sheet Excel export for L&D Stakeholders."""
    assert_batch_in_org(batch_id, db, current_user)
    batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == batch_id).first())
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    # 1. Gather Data
    groups = await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.batch_id == batch_id).all())
    group_ids = [g.id for g in groups]
    users = (
        await db.run_sync(lambda s: s.query(models.User)
        .filter(models.User.group_id.in_(group_ids), models.User.role == "Member")
        .all())
    )
    user_ids = [u.id for u in users]

    quiz_attempts = (
        await db.run_sync(lambda s: s.query(models.Attempt).filter(models.Attempt.user_id.in_(user_ids)).all())
    )
    coding_attempts = (
        await db.run_sync(lambda s: s.query(models.CodingAttempt)
        .filter(models.CodingAttempt.user_id.in_(user_ids))
        .all())
    )

    # 2. Setup Workbook
    wb = Workbook()

    # --- Sheet 1: Executive Insights ---
    ws1 = wb.active  # type: ignore
    assert ws1 is not None
    ws1.title = "Executive Insights"
    ws1.append(["GO-LIVE STRATEGIC ANALYSIS", "", "STUDYHUB V3 L&D ECOSYSTEM"])
    ws1.append(
        [
            f"Batch: {batch.name}",
            "",
            f"Generated: {datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M')}",
        ]
    )
    ws1.append([])

    # Header styling
    for cell in ws1[1]:
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )

    # Fetch AI Summary
    summary_data = {
        "groups": [g.name for g in groups],
        "user_count": len(users),
        "quiz_attempts": len(quiz_attempts),
        "coding_attempts": len(coding_attempts),
        "avg_quiz_score": sum(a.score for a in quiz_attempts) / len(quiz_attempts)
        if quiz_attempts
        else 0,
    }
    ai_bullets = await ai_executive.generate_batch_executive_summary(
        batch.name, summary_data
    )

    ws1.append(["AI-Generated Pedagogical Strategy:"])
    ai_data = ai_bullets.get("data", "")
    for i, line in enumerate(ai_data.split("\n")):
        if line.strip():
            ws1.append([line.strip()])
            ws1.cell(ws1.max_row, 1).alignment = Alignment(wrap_text=True)

    # --- Sheet 2: Performance Registry ---
    ws2 = wb.create_sheet("Performance Registry")
    ws2.append(
        [
            "User ID",
            "Full Name",
            "Group",
            "Department",
            "Avg Quiz Score (%)",
            "Avg Coding Score (10)",
            "Total Attempts",
        ]
    )

    user_map = {u.id: u for u in users}
    group_map = {g.id: g.name for g in groups}

    for uid, u in user_map.items():
        u_quizzes = [a for a in quiz_attempts if a.user_id == uid]
        u_coding = [a for a in coding_attempts if a.user_id == uid]

        avg_q = (
            (sum(a.score / a.total * 100 for a in u_quizzes) / len(u_quizzes))
            if u_quizzes
            else 0
        )
        avg_c = (sum(getattr(a, "score") or 0 for a in u_coding) / len(u_coding)) if u_coding else 0

        ws2.append(
            [
                uid,
                u.full_name,
                group_map.get(u.group_id, "N/A"),
                u.department_id,
                round(avg_q, 1),
                round(avg_c, 1),
                len(u_quizzes) + len(u_coding),
            ]
        )

    # --- Sheet 3: Raw Activity Log ---
    ws3 = wb.create_sheet("Raw Activity Log")
    ws3.append(
        ["Timestamp", "User", "Type", "Activity Name", "Score", "Total", "Status"]
    )

    # Combined logs
    logs = []
    for a in quiz_attempts:
        bank = (
            await db.run_sync(lambda s: s.query(models.QuestionBank)
            .filter(models.QuestionBank.id == a.bank_id)
            .first())
        )
        logs.append(
            [
                a.attempted_at,
                a.user_name,
                "Quiz",
                bank.name if bank else "Quiz",
                a.score,
                a.total,
                "N/A",
            ]
        )
    for a in coding_attempts:
        q = (
            await db.run_sync(lambda s: s.query(models.CodingQuestion)
            .filter(models.CodingQuestion.id == a.coding_question_id)
            .first())
        )
        logs.append(
            [
                a.attempted_at,
                user_map[a.user_id].full_name if a.user_id in user_map else "Unknown",
                "Coding",
                q.title if q else "Lab",
                a.score,
                10,
                "Verified" if a.is_verified else "Pending",
            ]
        )

    for log in sorted(logs, key=lambda x: x[0], reverse=True):
        ws3.append(log)

    # 3. Finalize
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=StudyHub_Batch_{batch_id}_Report.xlsx"
        },
    )

@router.get("/batch/{batch_id}/csv")
def export_batch_csv(
    batch_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_ldadmin),
):
    """CSV export for batch performance data — lighter alternative to Excel."""
    assert_batch_in_org(batch_id, db, current_user)
    import csv
    import io as _io

    from fastapi.responses import StreamingResponse as SR

    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    groups = db.query(models.Group).filter(models.Group.batch_id == batch_id).all()
    group_ids = [g.id for g in groups]
    users = db.query(models.User).filter(models.User.group_id.in_(group_ids)).all()
    user_ids = [u.id for u in users]

    quiz_attempts = (
        db.query(models.Attempt).filter(models.Attempt.user_id.in_(user_ids)).all()
    )
    coding_attempts = (
        db.query(models.CodingAttempt)
        .filter(models.CodingAttempt.user_id.in_(user_ids))
        .all()
    )

    group_map = {g.id: g.name for g in groups}

    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "User ID",
            "Full Name",
            "Group",
            "Email",
            "Avg Quiz Score (%)",
            "Avg Coding Score",
            "Total Attempts",
            "Coding Attempts",
        ]
    )

    for u in users:
        u_quiz = [a for a in quiz_attempts if a.user_id == u.id]
        u_code = [a for a in coding_attempts if a.user_id == u.id]
        avg_q = (
            round(sum(a.score / a.total * 100 for a in u_quiz) / len(u_quiz), 1)
            if u_quiz
            else 0
        )
        avg_c = round(sum(getattr(a, "score") or 0 for a in u_code) / len(u_code), 1) if u_code else 0
        writer.writerow(
            [
                u.id,
                u.full_name,
                group_map.get(u.group_id, "N/A"),
                u.email or "",
                avg_q,
                avg_c,
                len(u_quiz),
                len(u_code),
            ]
        )

    output.seek(0)
    return SR(
        iter([output.getvalue().encode()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=StudyHub_Batch_{batch_id}.csv"
        },
    )
